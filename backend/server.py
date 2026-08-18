"""
Backend de demo — Azzorti Benchmarking de Precios (Hito 1 / Retail).

Reemplaza el patrón "GitHub como base de datos" del prototipo original
(token embebido en el APK + capturas.json en un repo público) por una API
real con persistencia en SQLite. Pensado para correr en la laptop de
Yohana durante la demo remota: la app y el dashboard le apuntan a la IP
de la laptop en la red local, no hace falta hosting público.

Los datos de "azzorti_producto" y "politica_precio" son de MUESTRA para
la demo (no es el archivo SID real todavía) — quedan marcados como tal
en /info para que se pueda aclarar en la reunión qué es dato real y qué
es de muestra.

Ejecutar:
    pip install -r requirements.txt
    uvicorn server:app --host 0.0.0.0 --port 8000

Con eso queda accesible desde el celular en la misma WiFi en
http://<IP-de-la-laptop>:8000
"""
from __future__ import annotations

import base64
import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import io
import unicodedata
import zipfile
from xml.etree import ElementTree as ET

import fitz  # PyMuPDF - renderiza paginas de catalogos PDF para el OCR de ofertas
import openpyxl
import pytesseract
from PIL import Image
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

DB_PATH = Path(__file__).parent / "azzorti_demo.db"
STATIC_DIR = Path(__file__).parent / "static"
CAPTURAS_FOTOS_DIR = STATIC_DIR / "capturas_fotos"
CAPTURAS_FOTOS_DIR.mkdir(parents=True, exist_ok=True)
CATALOGOS_DIR = STATIC_DIR / "catalogos_competidor"
CATALOGOS_DIR.mkdir(parents=True, exist_ok=True)
ESTRELLA_FOTOS_DIR = STATIC_DIR / "productos_estrella_fotos"
ESTRELLA_FOTOS_DIR.mkdir(parents=True, exist_ok=True)
OFERTA_FOTOS_DIR = STATIC_DIR / "ofertas_fotos"
OFERTA_FOTOS_DIR.mkdir(parents=True, exist_ok=True)
CATALOGO_PAGINAS_DIR = STATIC_DIR / "catalogo_paginas"
CATALOGO_PAGINAS_DIR.mkdir(parents=True, exist_ok=True)

# En Windows el instalador de Tesseract no siempre queda en el PATH de la
# sesion actual - si "tesseract" no se encuentra por PATH, se prueba la
# ruta de instalacion por defecto de UB-Mannheim antes de fallar.
try:
    pytesseract.get_tesseract_version()
except Exception:
    _ruta_alterna = Path.home() / "AppData/Local/Programs/Tesseract-OCR/tesseract.exe"
    if _ruta_alterna.exists():
        pytesseract.pytesseract.tesseract_cmd = str(_ruta_alterna)

app = FastAPI(title="Azzorti Benchmarking — Backend de demo")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# La app Flutter y el dashboard HTML corren en orígenes distintos (o sin
# origen, desde un WebView/APK) — para la demo se permite cualquiera.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ====================== ESQUEMA Y SEED ======================

SCHEMA = """
CREATE TABLE IF NOT EXISTS captura (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    competidor TEXT NOT NULL,
    canal TEXT NOT NULL,
    campana TEXT NOT NULL,
    categoria TEXT NOT NULL,
    nivel_precio TEXT NOT NULL,
    descripcion TEXT,
    silueta TEXT,
    talla TEXT,
    composicion1 TEXT,
    composicion2 TEXT,
    manga TEXT,
    color TEXT,
    detalle TEXT,
    caracteristicas TEXT,
    precio REAL NOT NULL,
    sku_competidor TEXT,
    azzorti_sku_confirmado TEXT,
    foto_archivo TEXT,
    creada_en TEXT NOT NULL,
    UNIQUE (competidor, campana, sku_competidor)
);

-- Sin nivel_precio: el catalogo real de Azzorti no viene en niveles
-- Bajo/Medio/Alto (eso solo aplicaba a como Mercadeo agrupaba a la
-- competencia) - cada articulo real tiene un unico precio (P.O.).
-- descripcion/pagina_catalogo/foto_archivo: para que el analista pueda
-- confirmar visualmente la homologacion (no solo por texto).
CREATE TABLE IF NOT EXISTS azzorti_producto (
    sku TEXT PRIMARY KEY,
    categoria TEXT NOT NULL,
    descripcion TEXT,
    color TEXT,
    composicion TEXT,
    silueta TEXT,
    manga TEXT,
    precio REAL NOT NULL,
    campana TEXT NOT NULL,
    pagina_catalogo INTEGER,
    foto_archivo TEXT
);

-- politica_precio queda en el esquema por compatibilidad con datos viejos,
-- pero ya no se usa en la evaluacion (decision de Yohana: el umbral unico
-- de configuracion reemplaza las politicas puntuales por categoria, tanto
-- en Retail como en Venta Directa - ver Pendiente 3).
CREATE TABLE IF NOT EXISTS politica_precio (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    categoria TEXT NOT NULL,
    competidor TEXT NOT NULL,
    tipo TEXT NOT NULL,          -- DEBAJO_PCT | ENCIMA_PCT | IGUAL | SIN_POLITICA
    umbral_pct REAL,
    UNIQUE (categoria, competidor)
);

-- Configuracion editable desde el dashboard (umbral de alerta, tipo de
-- cambio, etc.) - clave/valor simple. El umbral aplica "hacia adelante"
-- (no se recalculan campanas historicas ya guardadas con otro umbral, tal
-- como se acordo para el Tipo de Cambio).
CREATE TABLE IF NOT EXISTS configuracion (
    clave TEXT PRIMARY KEY,
    valor TEXT NOT NULL,
    actualizado_en TEXT NOT NULL
);

-- Catalogo digital de un competidor de Venta Directa, subido por campana.
-- El analista lo sube una vez por campana y de ahi saca los "productos
-- estrella" (ver captura.catalogo_id) en vez de que el sistema intente
-- leer el catalogo completo solo (cada competidor trae un diseno de
-- catalogo distinto, no hay una sola forma automatica de leerlos todos).
CREATE TABLE IF NOT EXISTS catalogo_competidor (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    competidor TEXT NOT NULL,
    campana TEXT NOT NULL,
    archivo TEXT NOT NULL,
    subido_en TEXT NOT NULL
);

-- Lista fija de "productos estrella" definida por Mercadeo (no la decide
-- el motor de similitud): por cada producto del competidor, o ya viene con
-- un referente Azzorti especifico asignado a mano, o viene marcado para
-- compararse contra su propio precio de la campana anterior (cuando no
-- hay equivalente directo en el catalogo Azzorti). Se importa desde el
-- Excel real que maneja Yohana ("Comparativo fragancias azzorti vs
-- competencia.xlsx"), no se inventa la lista aqui.
CREATE TABLE IF NOT EXISTS producto_estrella (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    competidor TEXT NOT NULL,
    categoria TEXT,
    descripcion_competidor TEXT NOT NULL,
    modo TEXT NOT NULL,              -- HOMOLOGO_FIJO | VS_CAMPANA_ANTERIOR
    azzorti_referente TEXT,          -- descripcion del producto Azzorti (solo si modo=HOMOLOGO_FIJO)
    precio_competidor REAL,
    precio_azzorti REAL,
    foto_competidor TEXT,
    foto_azzorti TEXT,
    campana TEXT NOT NULL,
    -- Campaña que trae el archivo de Mercadeo del lado Azzorti (columna J):
    -- en HOMOLOGO_FIJO suele ser la misma campaña actual; en
    -- VS_CAMPANA_ANTERIOR es la campaña contra la que se comparo (hoy, C-09).
    campana_azzorti TEXT,
    -- % de variacion que Mercadeo ya calculo a mano en el archivo para las
    -- filas VS_CAMPANA_ANTERIOR (no hay precio Azzorti que calcular ahi).
    delta_pct_archivo REAL,
    actualizado_en TEXT NOT NULL,
    UNIQUE (competidor, descripcion_competidor, campana)
);

-- Set de referencia de logotipos/nombres de oferta por competidor, tal
-- como los define Mercadeo (importado desde "Detalle de ofertas.xlsx").
-- No se inventan nombres de oferta aqui: son el vocabulario contra el que
-- se compara el texto que el OCR lee en cada pagina de catalogo.
CREATE TABLE IF NOT EXISTS oferta_referencia (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    competidor TEXT NOT NULL,
    nombre_oferta TEXT NOT NULL,
    foto TEXT,
    actualizado_en TEXT NOT NULL,
    UNIQUE (competidor, nombre_oferta)
);

-- Resultado de escanear (OCR) las paginas de un catalogo_competidor en
-- busca de bandas de oferta. Se asigna por producto (via el "cod. NNNNN"
-- mas cercano a la banda detectada, ver _detectar_ofertas_en_pagina) -
-- producto_codigo queda NULL solo cuando la pagina no tiene ningun codigo
-- de producto legible por OCR (degradacion visible a nivel de pagina).
CREATE TABLE IF NOT EXISTS catalogo_oferta (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    catalogo_id INTEGER NOT NULL REFERENCES catalogo_competidor(id),
    pagina INTEGER NOT NULL,
    producto_codigo TEXT,
    oferta_detectada TEXT NOT NULL,
    score REAL,
    texto_ocr TEXT,
    creado_en TEXT NOT NULL,
    UNIQUE (catalogo_id, pagina, oferta_detectada, producto_codigo)
);

-- Indice (OCR) de los productos de un catalogo_competidor - se usa para
-- homologar Venta Directa contra el catalogo real de Azzorti subido en
-- Competencia (Pendiente app: antes solo Retail tenia con que homologar,
-- via azzorti_producto; Venta Directa no tenia ningun catalogo Azzorti
-- estructurado). texto_cercano es lo que el OCR leyo alrededor del "COD."
-- de cada producto (nombre, precio, descripcion corta) - se usa para
-- comparar por palabras clave, igual de aproximado que la deteccion de
-- ofertas, porque el catalogo real no tiene una fila por producto como
-- si tiene el Excel de Retail.
CREATE TABLE IF NOT EXISTS catalogo_producto (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    catalogo_id INTEGER NOT NULL REFERENCES catalogo_competidor(id),
    pagina INTEGER NOT NULL,
    producto_codigo TEXT NOT NULL,
    texto_cercano TEXT,
    precio REAL,
    seccion TEXT,
    creado_en TEXT NOT NULL,
    UNIQUE (catalogo_id, pagina, producto_codigo)
);
"""

# Reglas reales tomadas de "IPC BOLIVIA 2024_2025 - Venta retail.xlsx" y
# confirmadas por Yohana el 2026-07-27 (ver memoria del proyecto).
POLITICAS_SEED = [
    ("Blusas Femeninas", "Forever", "DEBAJO_PCT", 20),
    ("Vestidos Femeninos Cortos", "Forever", "DEBAJO_PCT", 30),
    ("Vestidos Femeninos Largos", "Forever", "DEBAJO_PCT", 30),
    ("Camisetas Masculinas", "Mitsuba", "IGUAL", None),
    ("Camisetas Femeninas", "Forever", "DEBAJO_PCT", 20),
    ("Camisetas Femeninas", "Mitsuba", "IGUAL", None),
    ("Jeans Femeninos", "Forever", "DEBAJO_PCT", 20),
    ("Jeans Femeninos", "Mitsuba", "DEBAJO_PCT", 20),
    ("Cubrecamas Dobles", "Casa In", "ENCIMA_PCT", 10),
    ("Sábanas Dobles", "Casa In", "ENCIMA_PCT", 10),
    ("Toallas", "Casa In", "ENCIMA_PCT", 10),
    ("Mochilas Infantiles", "Casa Ideas", "DEBAJO_PCT", 10),
    ("Cubrecamas Sencillos Infantiles", "Casa Ideas", "DEBAJO_PCT", 10),
]

# Tolerancia genérica de vigilancia. Valor de arranque/respaldo: 10%
# (confirmado por Yohana). Desde el Pendiente 3 este umbral reemplaza las
# políticas puntuales por categoría, y es editable en caliente via
# /configuracion/umbral-alerta - el valor real vigente se lee de la tabla
# "configuracion", esta constante solo es el valor con el que se siembra
# esa tabla la primera vez.
UMBRAL_ALERTA_GENERICA_PCT = 10.0


def _umbral_alerta_actual(conn) -> float:
    fila = conn.execute("SELECT valor FROM configuracion WHERE clave = 'umbral_alerta_pct'").fetchone()
    if fila is None:
        return UMBRAL_ALERTA_GENERICA_PCT
    try:
        return float(fila["valor"])
    except (TypeError, ValueError):
        return UMBRAL_ALERTA_GENERICA_PCT

# Catálogo Azzorti REAL de la campaña C-10 (202610), cruzado entre
# "inv_prca_202610.xlsx" (código, categoría, precio oficial) y el catálogo
# digital "BOL202610NAL.pdf" (composición, silueta y color leído de la
# foto de cada referencia). Cubre las categorías con política de precio
# que ya se pudieron verificar con ambas fuentes; el resto de categorías
# se agrega cuando se revisen sus páginas correspondientes del catálogo.
AZZORTI_PRODUCTOS_SEED = [
    # sku, categoria, descripcion, color, composicion, silueta, manga, precio, campana, pagina_catalogo, foto_archivo
    # Blusas Femeninas (subgrupo real "201 RE-BLUSAS FEM")
    ("R4874", "Blusas Femeninas", "Blusa Ref.R4874", "Verde", "Poliéster spandex acanalado", "Semiajustada", "Manga corta", 179.99, "C10 2026", 7, "catalogo_c10/pagina7.png"),
    ("R6844", "Blusas Femeninas", "Blusa Ref.R6844", None, "Tela tipo gamuza", "Semiajustada", "Manga corta", 209.99, "C10 2026", 17, "catalogo_c10/pagina17.png"),
    ("R6856", "Blusas Femeninas", "Blusa Ref.R6856", None, "Tejido doble punto pesado en poliéster algodón con spandex", "Semiajustada", "Manga corta", 189.99, "C10 2026", 18, "catalogo_c10/pagina18.png"),
    # Camisetas Femeninas (subgrupo real "207 RE-CAMISETAS FEM") — foto
    # recortada individual, no la página completa (las 3 comparten página
    # y mostrar la misma imagen para las 3 confundía la comparación). El
    # catálogo real llama a esto "Polera", no "Camiseta" — se corrige la
    # descripción para que coincida con el nombre real del producto.
    ("R4484", "Camisetas Femeninas", "Polera Ref.R4484", "Verde", "Poliéster algodón", "Semiajustada", "Manga corta", 139.99, "C10 2026", 24, "catalogo_c10/crop_R4484.png"),
    ("R6402", "Camisetas Femeninas", "Polera Ref.R6402", "Beige", "Poliéster algodón", "Semiajustada", "Manga corta", 139.99, "C10 2026", 24, "catalogo_c10/crop_R6402.png"),
    # SKU real = talla M (893460); "Betina" tiene 4 codigos, uno por talla
    # (S=277009, M=893460, L=269897, XL=678779) y la app no captura talla
    # todavia, asi que se usa M como representativo (ver nota pendiente).
    ("893460", "Camisetas Femeninas", "Polera Betina", "Blanco", "Poliéster algodón", "Semiajustada", "Manga corta", 139.99, "C10 2026", 24, "catalogo_c10/crop_Betina.png"),
    # Camisetas Masculinas (subgrupo real "207 RE-CAMISETAS MSC"). Foto
    # individual recortada (antes las 4 compartian la pagina completa, lo
    # que las hacia indistinguibles en pantalla). Pagina real impresa = 89,
    # no 91: el indice del PDF trae un desfase de 2 respecto al numero
    # impreso en el catalogo (mismo desfase encontrado en "Betina").
    ("R4808", "Camisetas Masculinas", "Polera Ref.R4808", "Naranja", "Algodón 100%", "Semiajustada", "Manga corta", 159.99, "C10 2026", 89, "catalogo_c10/crop_R4808.png"),
    ("R4326", "Camisetas Masculinas", "Polera Ref.R4326", "Amarillo", "Algodón 100%", "Semiajustada", "Manga corta", 159.99, "C10 2026", 89, "catalogo_c10/crop_R4326.png"),
    ("R4999", "Camisetas Masculinas", "Polera Ref.R4999", "Verde", "Algodón 100%", "Semiajustada", "Manga corta", 159.99, "C10 2026", 89, "catalogo_c10/crop_R4999.png"),
    ("R4813", "Camisetas Masculinas", "Polera Ref.R4813", "Verde claro", "Algodón 100%", "Semiajustada", "Manga corta", 159.99, "C10 2026", 89, "catalogo_c10/crop_R4813.png"),
    # Polos Masculinos (subgrupo real "Polos", pág. impresa 76) — categoria
    # nueva, distinta de "Camisetas Masculinas": son polos con cuello y
    # botones/cierre, precio real Bs 249.99 (mas caro que la polera lisa de
    # Bs 159.99). No existian en el catalogo de demo — se agregan porque una
    # captura real de un polo no tenia con que homologar bien (todo lo que
    # habia era poleras cuello redondo, similitud baja aunque fuera correcto).
    ("R6807", "Polos Masculinos", "Polo Ref.R6807", "Negro", "Poliéster algodón (piqué)", "Semiajustada", "Manga corta", 249.99, "C10 2026", 76, "catalogo_c10/crop_R6807.png"),
    ("R6808", "Polos Masculinos", "Polo Ref.R6808", "Blanco", "Poliéster (lanilla)", "Semiajustada", "Manga corta", 249.99, "C10 2026", 76, "catalogo_c10/crop_R6808.png"),
    # Crop Top Femenino (subgrupo real "233 RE-CROPTOP FEM", pág. 6) —
    # categoría que faltaba por completo; el sistema nunca podía sugerirla
    # porque la app tampoco la tenía como opción de categoría.
    ("R5323", "Crop Top Femenino", "Crop Top Ref.R5323", "Vino", "Tejido de punto piel de durazno poliéster spandex", "Ajustada", "Sin manga", 119.99, "C10 2026", 6, "catalogo_c10/crop_R5323.png"),
    ("RM1060", "Crop Top Femenino", "Crop Top Ref.RM1060", "Verde claro", "Tejido de punto piel de durazno poliéster spandex", "Ajustada", "Sin manga", 119.99, "C10 2026", 6, "catalogo_c10/crop_RM1060.png"),
    ("R5325", "Crop Top Femenino", "Crop Top Ref.R5325", "Verde", "Tejido de punto piel de durazno poliéster spandex", "Ajustada", "Sin manga", 119.99, "C10 2026", 6, "catalogo_c10/crop_R5325.png"),
    # Fragancias (Hito 2 / Venta Directa) — de "INDEX PRECIO SECTOR VENTA
    # DIRECTA1.Manual.xlsx", hoja "Azzorti vs Natura y Esika", sección
    # AZZORTI (precio real del ciclo C05). Foto real extraída de la misma
    # hoja (imagen "en celda" de Excel — se rastreó celda → vm → metadata
    # → richValue → richValueRel → archivo real en xl/media/). Sin
    # color/silueta/manga: el motor de similitud se diseñó para ropa: para
    # fragancias, la comparación se apoya más en el nombre del producto.
    # SKU: codigo real de producto Azzorti (no un identificador inventado).
    # Confirmados contra "CATALOGO BOL C-5-2026 baja.pdf" (catalogo real de
    # Venta Directa, campana C-5 2026): STROM=277983 (ya estaba bien),
    # SHORT DISTANCE=755712, PRIM ROSE AMA=977213 (nombre real completo, no
    # solo "Primrose"; precio tambien corregido a 219.99). Virtuossa NO
    # aparece en este catalogo (busqueda exhaustiva de texto) - puede que ya
    # no este en esta campana; codigo sigue pendiente de confirmar con Yohana.
    ("VIRTUOSSA", "Fragancias", "Virtuossa 60 ML", None, None, None, None, 219.99, "C05 2026", None, "venta_directa/virtuossa.png"),
    ("977213", "Fragancias", "Prim Rose Ama 50 ML", None, None, None, None, 219.99, "C05 2026", None, "venta_directa/primrose.png"),
    ("755712", "Fragancias", "Short Distance 80 ML", None, None, None, None, 179.99, "C05 2026", None, "venta_directa/shortdistance.png"),
    ("277983", "Fragancias", "Strom Poder Energía", None, None, None, None, 229.99, "C05 2026", None, "venta_directa/strom.png"),
]


@contextmanager
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    with get_conn() as conn:
        conn.executescript(SCHEMA)
        # "CREATE TABLE IF NOT EXISTS" no agrega columnas nuevas a una tabla
        # que ya existia de antes (la base de datos de la demo no se borra
        # en cada reinicio) - se agrega la columna a mano si falta.
        try:
            conn.execute("ALTER TABLE captura ADD COLUMN talla TEXT")
        except sqlite3.OperationalError:
            pass  # ya existe
        try:
            conn.execute("ALTER TABLE captura ADD COLUMN catalogo_id INTEGER")
        except sqlite3.OperationalError:
            pass  # ya existe
        try:
            conn.execute("ALTER TABLE producto_estrella ADD COLUMN foto_competidor TEXT")
        except sqlite3.OperationalError:
            pass  # ya existe
        try:
            conn.execute("ALTER TABLE producto_estrella ADD COLUMN foto_azzorti TEXT")
        except sqlite3.OperationalError:
            pass  # ya existe
        try:
            conn.execute("ALTER TABLE producto_estrella ADD COLUMN campana_azzorti TEXT")
        except sqlite3.OperationalError:
            pass  # ya existe
        try:
            conn.execute("ALTER TABLE producto_estrella ADD COLUMN delta_pct_archivo REAL")
        except sqlite3.OperationalError:
            pass  # ya existe
        # catalogo_oferta es cache de OCR (se regenera con POST
        # /catalogos/{id}/detectar-ofertas) - si quedo de una version
        # anterior sin la columna producto_codigo (deteccion por pagina
        # completa, ver Pendiente 8 v1), se recrea con el esquema nuevo
        # en vez de intentar migrar datos derivados que no vale la pena
        # conservar.
        columnas = {r["name"] for r in conn.execute("PRAGMA table_info(catalogo_oferta)")}
        if columnas and "producto_codigo" not in columnas:
            conn.execute("DROP TABLE catalogo_oferta")
            conn.executescript(SCHEMA)
        # catalogo_producto tambien es cache de OCR (se regenera con POST
        # /catalogos/{id}/indexar-productos) - se recrea igual si le falta
        # la columna 'seccion' (agregada para poder filtrar candidatos de
        # homologacion por categoria en vez de comparar contra todo el
        # catalogo mezclado).
        columnas_prod = {r["name"] for r in conn.execute("PRAGMA table_info(catalogo_producto)")}
        if columnas_prod and "seccion" not in columnas_prod:
            conn.execute("DROP TABLE catalogo_producto")
            conn.executescript(SCHEMA)
        # INSERT OR IGNORE por fila (no "solo si la tabla está vacía"): la
        # base de datos ya no se borra en cada reinicio, así que hay que
        # poder agregar productos/políticas nuevas al seed sin duplicar ni
        # perder lo que el analista ya capturó.
        conn.executemany(
            "INSERT OR IGNORE INTO politica_precio (categoria, competidor, tipo, umbral_pct) VALUES (?, ?, ?, ?)",
            POLITICAS_SEED,
        )
        conn.executemany(
            "INSERT OR IGNORE INTO azzorti_producto (sku, categoria, descripcion, color, composicion, silueta, manga, precio, campana, pagina_catalogo, foto_archivo) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            AZZORTI_PRODUCTOS_SEED,
        )
        conn.execute(
            "INSERT OR IGNORE INTO configuracion (clave, valor, actualizado_en) VALUES ('umbral_alerta_pct', ?, ?)",
            (str(UMBRAL_ALERTA_GENERICA_PCT), datetime.now(timezone.utc).isoformat()),
        )


init_db()


# ====================== MODELOS ======================

class CapturaIn(BaseModel):
    competidor: str
    canal: str = "Retail"
    campana: str
    categoria: str
    nivel_precio: str
    descripcion: Optional[str] = None
    silueta: Optional[str] = None
    talla: Optional[str] = None
    composicion1: Optional[str] = None
    composicion2: Optional[str] = None
    manga: Optional[str] = None
    color: Optional[str] = None
    detalle: Optional[str] = None
    caracteristicas: Optional[str] = None
    precio: float
    sku_competidor: Optional[str] = None
    foto_producto_b64: Optional[str] = None
    catalogo_id: Optional[int] = None


class HomologacionConfirmar(BaseModel):
    azzorti_sku: str


# ====================== ENDPOINTS ======================

@app.get("/info")
def info():
    with get_conn() as conn:
        umbral = _umbral_alerta_actual(conn)
    return {
        "servicio": "Azzorti Benchmarking — backend de demo",
        "nota": (
            "azzorti_producto (campana C-10) es dato real de negocio, cruzado "
            "entre inv_prca_202610.xlsx y el catalogo digital BOL202610NAL.pdf. "
            "politica_precio quedo retirada de la evaluacion (Pendiente 3): el "
            "umbral unico de configuracion reemplaza las politicas puntuales "
            "por categoria, tanto en Retail como en Venta Directa."
        ),
        "umbral_alerta_generica_pct": umbral,
    }


class UmbralAlertaIn(BaseModel):
    umbral_pct: float


class TipoCambioIn(BaseModel):
    tc: float


@app.get("/configuracion/umbral-alerta")
def obtener_umbral_alerta():
    with get_conn() as conn:
        return {"umbral_pct": _umbral_alerta_actual(conn)}


@app.post("/configuracion/umbral-alerta")
def actualizar_umbral_alerta(body: UmbralAlertaIn):
    """Actualiza el umbral de alerta unico (reemplaza las politicas por
    categoria en ambos canales, ver Pendiente 3). Aplica hacia adelante:
    las evaluaciones ya calculadas y guardadas (ninguna se guarda hoy, se
    calculan al vuelo) no se tocan retroactivamente porque este valor solo
    se lee al momento de pedir una evaluacion nueva - la campana activa en
    ese momento es la que usa el valor nuevo."""
    if body.umbral_pct <= 0:
        raise HTTPException(400, "El umbral debe ser un porcentaje mayor a 0.")
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO configuracion (clave, valor, actualizado_en) VALUES ('umbral_alerta_pct', ?, ?) "
            "ON CONFLICT (clave) DO UPDATE SET valor=excluded.valor, actualizado_en=excluded.actualizado_en",
            (str(body.umbral_pct), datetime.now(timezone.utc).isoformat()),
        )
    return {"umbral_pct": body.umbral_pct, "mensaje": "Umbral actualizado."}


@app.get("/configuracion/tipo-cambio")
def obtener_tipo_cambio():
    with get_conn() as conn:
        fila = conn.execute("SELECT valor FROM configuracion WHERE clave = 'tipo_cambio'").fetchone()
    return {"tc": float(fila["valor"]) if fila else 6.96}


@app.post("/configuracion/tipo-cambio")
def actualizar_tipo_cambio(body: TipoCambioIn):
    """Mismo patron que el umbral de alerta (Pendiente 3): un solo valor
    editable desde el dashboard, guardado en 'configuracion'. Aplica hacia
    adelante - no recalcula conversiones ya mostradas de campanas pasadas."""
    if body.tc <= 0:
        raise HTTPException(400, "El tipo de cambio debe ser mayor a 0.")
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO configuracion (clave, valor, actualizado_en) VALUES ('tipo_cambio', ?, ?) "
            "ON CONFLICT (clave) DO UPDATE SET valor=excluded.valor, actualizado_en=excluded.actualizado_en",
            (str(body.tc), datetime.now(timezone.utc).isoformat()),
        )
    return {"tc": body.tc, "mensaje": "Tipo de cambio actualizado."}


def _guardar_foto_captura(competidor: str, foto_b64: str) -> str:
    """Decodifica y guarda la foto del producto capturada en el celular.
    Devuelve solo el nombre de archivo (se sirve luego desde /static)."""
    slug = re.sub(r"[^a-z0-9]+", "-", competidor.lower()).strip("-") or "captura"
    nombre = f"{slug}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S%f')}.jpg"
    (CAPTURAS_FOTOS_DIR / nombre).write_bytes(base64.b64decode(foto_b64))
    return nombre


@app.post("/catalogos", status_code=201)
async def subir_catalogo(
    competidor: str = Form(...),
    campana: str = Form(...),
    archivo: UploadFile = File(...),
):
    """Sube el catalogo digital de un competidor de Venta Directa para una
    campana. No se intenta leer/extraer nada del archivo automaticamente
    (cada competidor trae un diseno distinto) - queda disponible para que
    el analista lo revise y registre a mano los "productos estrella" que
    Mercadeo ya identifico, referenciando este catalogo (ver catalogo_id
    en POST /capturas)."""
    slug = re.sub(r"[^a-z0-9]+", "-", competidor.lower()).strip("-") or "competidor"
    extension = Path(archivo.filename or "").suffix or ".pdf"
    nombre = f"{slug}_{re.sub(r'[^a-z0-9]+', '-', campana.lower())}_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')}{extension}"
    # Antes usaba archivo.file.read() (sincrono/bloqueante) - con un PDF
    # grande (20+ MB) eso congelaba todo el backend (un solo hilo async)
    # mientras se leia, incluso otras peticiones dejaban de responder.
    # await archivo.read() no bloquea el resto del servidor.
    contenido = await archivo.read()
    (CATALOGOS_DIR / nombre).write_bytes(contenido)
    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO catalogo_competidor (competidor, campana, archivo, subido_en) VALUES (?, ?, ?, ?)",
            (competidor, campana, nombre, datetime.now(timezone.utc).isoformat()),
        )
        catalogo_id = cur.lastrowid
    return {"id": catalogo_id, "mensaje": "Catálogo subido correctamente."}


@app.get("/catalogos")
def listar_catalogos(request: Request, competidor: Optional[str] = None, campana: Optional[str] = None):
    query = "SELECT * FROM catalogo_competidor WHERE 1=1"
    params: list = []
    if competidor:
        query += " AND competidor = ?"
        params.append(competidor)
    if campana:
        query += " AND campana = ?"
        params.append(campana)
    query += " ORDER BY id DESC"
    with get_conn() as conn:
        rows = conn.execute(query, params).fetchall()
    return [
        {**dict(r), "archivo_url": f"{request.base_url}static/catalogos_competidor/{r['archivo']}"}
        for r in rows
    ]


def _texto(valor) -> str:
    """Convierte a texto plano y limpia espacios/None - los errores de
    formula (#VALUE!, etc.) llegan como objetos que no son str, se tratan
    como vacios en vez de reventar la importacion."""
    if valor is None:
        return ""
    if not isinstance(valor, str):
        return ""
    return valor.strip()


# Formato estandar de campaña en todo el sistema: "C10 2026" (2 digitos +
# año) - los archivos viejos traian "C-10" (sin año, con guion).
_CAMPANA_ESTANDAR_RE = re.compile(r"^C\d{2} \d{4}$")


def _numero(valor):
    """Solo acepta numeros reales - una celda con #VALUE! o vacia se
    guarda como None (precio aun no cargado), nunca se inventa un 0."""
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return float(valor)
    return None


def _parsear_productos_estrella(contenido: bytes) -> list[dict]:
    """Lee el Excel real de 'productos estrella' que maneja Mercadeo (ver
    memoria del proyecto: 'Comparativo fragancias azzorti vs competencia').
    El archivo real repite encabezados iguales dos veces ("PRECIO", "FOTO",
    "CAMPAÑA" aparecen una vez para el lado del competidor y otra para el
    lado Azzorti) - buscarlos solo por texto los confundia (el ultimo que
    calzaba se quedaba con la columna, pisando al primero). Por eso, salvo
    "descripcion" y "referente" (las unicas dos columnas con texto que no
    se repite), el resto de columnas se ubican por POSICION relativa a esas
    dos, que es fija en el archivo real: Competidor(A) Categoria(B)
    Descripcion(C) Foto(D) Precio(E) Campaña(F) Referente(G) Foto(H)
    Precio(I) Campaña(J)."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    filas: list[dict] = []
    for ws in wb.worksheets:
        # Busca la fila de encabezado: la primera que tenga una celda que
        # empiece con "descri" (cubre "DESCRICION"/"DESCRIPCION").
        fila_encabezado = None
        for r in range(1, min(ws.max_row, 15) + 1):
            for c in range(1, ws.max_column + 1):
                if _texto(ws.cell(r, c).value).lower().startswith("descri"):
                    fila_encabezado = r
                    break
            if fila_encabezado:
                break
        if not fila_encabezado:
            continue  # esta hoja no tiene la tabla esperada

        col_descripcion = None
        col_referente = None
        for c in range(1, ws.max_column + 1):
            texto = _texto(ws.cell(fila_encabezado, c).value).lower()
            if texto.startswith("descri"):
                col_descripcion = c
            elif ("referente" in texto or "azzorti" in texto) and "precio" not in texto:
                col_referente = c
        if not col_descripcion or not col_referente:
            continue  # encabezados que no calzan con el formato esperado

        col_competidor = 1  # columna A - siempre el competidor en el archivo real
        col_categoria = col_descripcion - 1
        col_foto = col_descripcion + 1
        col_precio_competidor = col_descripcion + 2
        col_campana_competidor = col_descripcion + 3
        col_foto_azzorti = col_referente + 1
        col_precio_azzorti = col_referente + 2
        col_campana_azzorti = col_referente + 3
        tiene_bloque_azzorti = col_campana_azzorti <= ws.max_column

        ultimo_competidor = None
        for r in range(fila_encabezado + 1, ws.max_row + 1):
            competidor = _texto(ws.cell(r, col_competidor).value) or ultimo_competidor
            if _texto(ws.cell(r, col_competidor).value):
                ultimo_competidor = competidor
            descripcion = _texto(ws.cell(r, col_descripcion).value)
            referente = _texto(ws.cell(r, col_referente).value)
            if not competidor or not descripcion or not referente:
                continue  # fila vacia/incompleta - se ignora, no se inventa nada

            es_vs_anterior = "campaña anterior" in referente.lower() or "campana anterior" in referente.lower()
            valor_col_azzorti = _numero(ws.cell(r, col_precio_azzorti).value) if tiene_bloque_azzorti else None
            filas.append({
                "competidor": competidor,
                "categoria": _texto(ws.cell(r, col_categoria).value) if col_categoria >= 1 else None,
                "descripcion_competidor": descripcion,
                "modo": "VS_CAMPANA_ANTERIOR" if es_vs_anterior else "HOMOLOGO_FIJO",
                # Se guarda el texto tal cual viene en el archivo en los dos
                # modos - en VS_CAMPANA_ANTERIOR no es un producto puntual,
                # pero es la explicación real de Mercadeo (ej. "comparar con
                # la campaña anterior, fragancias de un punto de precio
                # mayor al de azzorti") y no hay que reemplazarla por un
                # texto genérico nuestro.
                "azzorti_referente": referente,
                "precio_competidor": _numero(ws.cell(r, col_precio_competidor).value),
                "campana_competidor": _texto(ws.cell(r, col_campana_competidor).value) or None,
                # La columna se llama "Precio" en el archivo en los dos
                # modos - se respeta como precio en los dos y el % se
                # calcula en el sistema comparandolo contra el precio del
                # competidor, no se toma un % ya calculado del archivo.
                "precio_azzorti": valor_col_azzorti,
                # Algunos archivos traen esta columna en formato viejo
                # ("C-10", sin año) en vez del estandar ("C10 2026") - se
                # descarta si no calza con el estandar en vez de guardar un
                # valor inconsistente; el llamador usa la campaña del
                # formulario como respaldo (ver /productos-estrella/importar).
                "campana_azzorti": (
                    _texto(ws.cell(r, col_campana_azzorti).value)
                    if tiene_bloque_azzorti and _CAMPANA_ESTANDAR_RE.match(_texto(ws.cell(r, col_campana_azzorti).value))
                    else None
                ),
                "_fila_excel": r,
                "_col_foto": col_foto,
                "_col_foto_azzorti": col_foto_azzorti,
            })
    return filas


_XDR_NS = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
_A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
_R_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def _col_letras_a_numero(letras: str) -> int:
    n = 0
    for ch in letras:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _ruta_hoja_por_nombre(z: "zipfile.ZipFile", nombre_hoja: str) -> Optional[str]:
    """Resuelve 'NATURA ' -> 'xl/worksheets/sheet1.xml' (o el numero que
    corresponda). Necesario para libros con varias hojas (cada una trae su
    propio drawing/richvalue independiente) - sin esto, dos hojas distintas
    con una imagen en la misma fila se pisarian entre si."""
    nombres = set(z.namelist())
    if "xl/workbook.xml" not in nombres or "xl/_rels/workbook.xml.rels" not in nombres:
        return None
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    sheets = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml)
    rid = next((r for n, r in sheets if n == nombre_hoja), None)
    if not rid:
        return None
    rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid_to_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))
    target = rid_to_target.get(rid)
    return f"xl/{target}" if target else None


def _drawing_de_hoja(z: "zipfile.ZipFile", sheet_path: str) -> Optional[str]:
    """'xl/worksheets/sheet2.xml' -> 'xl/drawings/drawing2.xml' (la que esa
    hoja especifica declara en su propio archivo de relaciones)."""
    nombres = set(z.namelist())
    partes = sheet_path.rsplit("/", 1)
    rels_path = f"{partes[0]}/_rels/{partes[1]}.rels"
    if rels_path not in nombres:
        return None
    rels_xml = z.read(rels_path).decode("utf-8")
    for rel_id, target in re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml):
        if "drawing" in target and target.endswith(".xml"):
            return "xl/" + target.replace("../", "")
    return None


def _extraer_fotos_por_fila(contenido: bytes, nombre_hoja: Optional[str] = None) -> dict[int, bytes]:
    """Devuelve {fila_excel_1based: bytes_de_imagen} para fotos insertadas
    como dibujo flotante (forma anclada cerca de una celda - el mecanismo
    clasico de "Insertar imagen"). Si una fila tiene mas de una foto
    ancladas, se queda con la primera. No filtra por columna: sirve como
    respaldo para cuando la foto del competidor no vino por el mecanismo
    de "imagen en celda" (ver _extraer_fotos_por_celda).

    Si se pasa nombre_hoja, solo se lee el drawing propio de esa hoja - en
    libros con varias hojas (una por competidor), cada hoja tiene su propio
    archivo de dibujo con anclajes que reinician en la fila 0, asi que sin
    este filtro la foto de la fila 4 de una hoja se mezclaria con la de la
    fila 4 de otra."""
    fotos: dict[int, bytes] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(contenido)) as z:
            nombres = set(z.namelist())
            if nombre_hoja is not None:
                sheet_path = _ruta_hoja_por_nombre(z, nombre_hoja)
                drawing_only = _drawing_de_hoja(z, sheet_path) if sheet_path else None
                drawing_paths = [drawing_only] if drawing_only else []
            else:
                drawing_paths = sorted(n for n in nombres if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n))
            for drawing_path in drawing_paths:
                rels_path = drawing_path.replace("drawings/", "drawings/_rels/") + ".rels"
                if rels_path not in nombres:
                    continue
                rel_map = {
                    rel.get("Id"): rel.get("Target")
                    for rel in ET.fromstring(z.read(rels_path))
                }
                drawing_xml = ET.fromstring(z.read(drawing_path))
                anchors = drawing_xml.findall(f"{_XDR_NS}twoCellAnchor") + drawing_xml.findall(f"{_XDR_NS}oneCellAnchor")
                for anchor in anchors:
                    from_el = anchor.find(f"{_XDR_NS}from")
                    if from_el is None:
                        continue
                    fila_xml = from_el.find(f"{_XDR_NS}row")
                    if fila_xml is None or fila_xml.text is None:
                        continue
                    fila_excel = int(fila_xml.text) + 1  # xdr es 0-index, Excel es 1-index
                    if fila_excel in fotos:
                        continue
                    blip = anchor.find(f".//{_A_NS}blip")
                    if blip is None:
                        continue
                    embed_id = blip.get(f"{_R_NS}embed")
                    target = rel_map.get(embed_id)
                    if not target:
                        continue
                    media_path = "xl/" + target.replace("../", "")
                    if media_path in nombres:
                        fotos[fila_excel] = z.read(media_path)
    except Exception:
        pass  # best-effort: si algo falla, se importa sin fotos en vez de tumbar la importacion
    return fotos


def _extraer_fotos_ancladas_por_celda(contenido: bytes, nombre_hoja: Optional[str] = None) -> dict[tuple[int, int], bytes]:
    """Devuelve {(fila, columna) 1-based: bytes_de_imagen} para dibujos
    flotantes (twoCellAnchor/oneCellAnchor) que SI traen columna en su
    anclaje <xdr:from><xdr:col> - a diferencia de _extraer_fotos_por_fila
    (que ignora la columna a proposito y se queda solo con la primera
    imagen de cada fila, pensado para el caso de una sola foto por fila),
    esta función distingue 2 fotos ancladas a la MISMA fila pero en
    columnas distintas (ej. foto del competidor en columna D y foto de
    Azzorti en columna H de la misma fila) - antes la segunda se perdia
    silenciosamente."""
    fotos: dict[tuple[int, int], bytes] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(contenido)) as z:
            nombres = set(z.namelist())
            if nombre_hoja is not None:
                sheet_path = _ruta_hoja_por_nombre(z, nombre_hoja)
                drawing_only = _drawing_de_hoja(z, sheet_path) if sheet_path else None
                drawing_paths = [drawing_only] if drawing_only else []
            else:
                drawing_paths = sorted(n for n in nombres if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n))
            for drawing_path in drawing_paths:
                rels_path = drawing_path.replace("drawings/", "drawings/_rels/") + ".rels"
                if rels_path not in nombres:
                    continue
                rel_map = {
                    rel.get("Id"): rel.get("Target")
                    for rel in ET.fromstring(z.read(rels_path))
                }
                drawing_xml = ET.fromstring(z.read(drawing_path))
                anchors = drawing_xml.findall(f"{_XDR_NS}twoCellAnchor") + drawing_xml.findall(f"{_XDR_NS}oneCellAnchor")
                for anchor in anchors:
                    from_el = anchor.find(f"{_XDR_NS}from")
                    if from_el is None:
                        continue
                    fila_xml = from_el.find(f"{_XDR_NS}row")
                    col_xml = from_el.find(f"{_XDR_NS}col")
                    if fila_xml is None or fila_xml.text is None or col_xml is None or col_xml.text is None:
                        continue
                    fila_excel = int(fila_xml.text) + 1
                    col_excel = int(col_xml.text) + 1  # xdr es 0-index, Excel es 1-index
                    if (fila_excel, col_excel) in fotos:
                        continue
                    blip = anchor.find(f".//{_A_NS}blip")
                    if blip is None:
                        continue
                    embed_id = blip.get(f"{_R_NS}embed")
                    target = rel_map.get(embed_id)
                    if not target:
                        continue
                    media_path = "xl/" + target.replace("../", "")
                    if media_path in nombres:
                        fotos[(fila_excel, col_excel)] = z.read(media_path)
    except Exception:
        pass  # best-effort: si algo falla, se importa sin estas fotos en vez de tumbar la importacion
    return fotos


def _extraer_fotos_por_celda(contenido: bytes, nombre_hoja: Optional[str] = None) -> dict[tuple[int, int], bytes]:
    """Devuelve {(fila, columna) 1-based: bytes_de_imagen} para fotos
    insertadas como "imagen en celda" (rich value de Excel, distinto al
    dibujo flotante) - la misma tecnica ya usada antes para sacar las fotos
    de Virtuossa/Prim Rose Ama/Strom del excel original de Venta Directa:
    celda con atributo vm=N -> metadata.xml (indice N-1) -> rdrichvalue.xml
    -> richValueRel.xml -> richValueRel.xml.rels -> archivo real en media/.
    Se sigue la cadena completa en vez de asumir que vm=N == imageN.png,
    porque esa relacion es una coincidencia de este archivo, no una regla.

    Si se pasa nombre_hoja, solo se buscan celdas vm= dentro de esa hoja -
    metadata.xml/rdrichvalue.xml/richValueRel.xml son compartidos por todo
    el libro (esos si aplican a cualquier hoja por igual), pero las celdas
    vm= de cada hoja son propias de esa hoja.

    Ademas de "imagen en celda" (rich value), tambien se suman aqui los
    dibujos flotantes que SI tienen columna en su anclaje (ver
    _extraer_fotos_ancladas_por_celda) - un mismo archivo puede tener la
    foto del competidor como dibujo flotante y la de Azzorti tambien como
    dibujo flotante en otra columna (no siempre es "rich value"), y antes
    solo se buscaba la tecnica rich value aqui, dejando la de dibujo
    flotante sin distinguir columna (ver bug real: Estrella.xlsx trae las
    2 fotos como dibujo con columna, y la de Azzorti se perdia)."""
    fotos: dict[tuple[int, int], bytes] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(contenido)) as z:
            nombres = set(z.namelist())
            if "xl/metadata.xml" in nombres and "xl/richData/rdrichvalue.xml" in nombres:
                meta_xml = z.read("xl/metadata.xml").decode("utf-8")
                bk_list = re.findall(r"<bk>.*?<xlrd:rvb i=\"(\d+)\"/>.*?</bk>", meta_xml, re.DOTALL)
                rv_xml = z.read("xl/richData/rdrichvalue.xml").decode("utf-8")
                rv_list = re.findall(r"<rv[^>]*><v>(\d+)</v>", rv_xml)
                rel_rel_path = "xl/richData/richValueRel.xml"
                rels_path = "xl/richData/_rels/richValueRel.xml.rels"
                if rel_rel_path in nombres and rels_path in nombres:
                    rel_ids = re.findall(r'r:id="(rId\d+)"', z.read(rel_rel_path).decode("utf-8"))
                    rid_to_media = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"', z.read(rels_path).decode("utf-8")))

                    if nombre_hoja is not None:
                        sheet_path = _ruta_hoja_por_nombre(z, nombre_hoja)
                        sheet_paths = [sheet_path] if sheet_path else []
                    else:
                        sheet_paths = sorted(n for n in nombres if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", n))
                    for sheet_path in sheet_paths:
                        sheet_xml = z.read(sheet_path).decode("utf-8")
                        for letras, fila_txt, vm in re.findall(r'<c r="([A-Z]+)(\d+)"[^>]*vm="(\d+)"', sheet_xml):
                            try:
                                bk_idx = int(vm) - 1
                                rvb_i = int(bk_list[bk_idx])
                                rel_idx = int(rv_list[rvb_i])
                                rid = rel_ids[rel_idx]
                                media = rid_to_media.get(rid)
                            except (IndexError, ValueError):
                                continue
                            if not media:
                                continue
                            media_path = "xl/media/" + media
                            if media_path in nombres:
                                fotos[(int(fila_txt), _col_letras_a_numero(letras))] = z.read(media_path)
    except Exception:
        pass  # best-effort: si algo falla, se importa sin estas fotos en vez de tumbar la importacion
    # Se suman los dibujos flotantes con columna - sin pisar un hit real de
    # "rich value" si por algun motivo coincidieran en la misma celda.
    for clave, bytes_img in _extraer_fotos_ancladas_por_celda(contenido, nombre_hoja).items():
        fotos.setdefault(clave, bytes_img)
    return fotos


@app.post("/productos-estrella/importar")
async def importar_productos_estrella(
    campana: str = Form(...),
    archivo: UploadFile = File(...),
):
    """Importa/actualiza la lista de productos estrella desde el Excel real
    de Mercadeo para una campana. Re-importar el mismo archivo actualiza
    los precios (no duplica), gracias al UNIQUE(competidor, descripcion,
    campana)."""
    contenido = await archivo.read()
    try:
        filas = _parsear_productos_estrella(contenido)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el Excel: {e}")
    if not filas:
        raise HTTPException(
            400,
            "No se encontraron filas con el formato esperado (columnas "
            "Descripción y Referente Azzorti). Revisa los encabezados del archivo.",
        )
    fotos_por_fila = _extraer_fotos_por_fila(contenido)      # dibujo flotante (competidor, respaldo)
    fotos_por_celda = _extraer_fotos_por_celda(contenido)     # imagen en celda (competidor y/o Azzorti)

    def _guardar(foto_bytes: bytes, prefijo: str) -> str:
        extension = ".jpg" if foto_bytes[:3] == b"\xff\xd8\xff" else ".png"
        slug = re.sub(r"[^a-z0-9]+", "-", prefijo.lower()).strip("-")
        nombre = f"{slug}{extension}"
        (ESTRELLA_FOTOS_DIR / nombre).write_bytes(foto_bytes)
        return nombre

    con_foto_competidor = 0
    con_foto_azzorti = 0
    sospechosos: list[str] = []
    with get_conn() as conn:
        for f in filas:
            # Aviso, no bloqueo: hay filas donde la columna "Precio" del
            # lado Azzorti trae un numero muy chico (ej. 2.86) comparado con
            # el precio del competidor - se avisa igual, respetando el
            # valor tal cual venga en el archivo.
            if (
                f["precio_azzorti"] is not None
                and f["precio_competidor"]
                and f["precio_azzorti"] < f["precio_competidor"] * 0.2
            ):
                sospechosos.append(f"{f['competidor']} · {f['descripcion_competidor']} (Bs {f['precio_azzorti']})")
            fila_excel = f["_fila_excel"]
            col_foto = f["_col_foto"]
            col_foto_azzorti = f["_col_foto_azzorti"]

            foto_competidor_bytes = (
                (fotos_por_celda.get((fila_excel, col_foto)) if col_foto else None)
                or fotos_por_fila.get(fila_excel)
            )
            foto_competidor_nombre = None
            if foto_competidor_bytes:
                foto_competidor_nombre = _guardar(foto_competidor_bytes, f"{f['competidor']}-{f['descripcion_competidor']}")
                con_foto_competidor += 1

            foto_azzorti_bytes = fotos_por_celda.get((fila_excel, col_foto_azzorti))
            foto_azzorti_nombre = None
            if foto_azzorti_bytes:
                foto_azzorti_nombre = _guardar(foto_azzorti_bytes, f"azzorti-{f['azzorti_referente'] or f['descripcion_competidor']}")
                con_foto_azzorti += 1

            conn.execute(
                """INSERT INTO producto_estrella
                (competidor, categoria, descripcion_competidor, modo, azzorti_referente,
                 precio_competidor, precio_azzorti, campana_azzorti,
                 foto_competidor, foto_azzorti, campana, actualizado_en)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (competidor, descripcion_competidor, campana) DO UPDATE SET
                    categoria=excluded.categoria, modo=excluded.modo,
                    azzorti_referente=excluded.azzorti_referente,
                    precio_competidor=excluded.precio_competidor,
                    precio_azzorti=excluded.precio_azzorti,
                    campana_azzorti=excluded.campana_azzorti,
                    foto_competidor=COALESCE(excluded.foto_competidor, producto_estrella.foto_competidor),
                    foto_azzorti=COALESCE(excluded.foto_azzorti, producto_estrella.foto_azzorti),
                    actualizado_en=excluded.actualizado_en""",
                (
                    f["competidor"], f["categoria"], f["descripcion_competidor"], f["modo"],
                    f["azzorti_referente"], f["precio_competidor"], f["precio_azzorti"],
                    # Sin campana_azzorti valida en el archivo: si es
                    # HOMOLOGO_FIJO se asume que compara contra el catalogo
                    # Azzorti de la campaña actual (la del formulario) - en
                    # VS_CAMPANA_ANTERIOR no se asume nada, queda vacio.
                    f["campana_azzorti"] or (campana if f["modo"] == "HOMOLOGO_FIJO" else None),
                    foto_competidor_nombre, foto_azzorti_nombre,
                    campana, datetime.now(timezone.utc).isoformat(),
                ),
            )
    mensaje = (
        f"{len(filas)} productos estrella importados/actualizados "
        f"({con_foto_competidor} con foto de competidor, {con_foto_azzorti} con foto Azzorti)."
    )
    if sospechosos:
        mensaje += (
            f" ⚠ {len(sospechosos)} con precio Azzorti sospechosamente bajo, revisa esa columna en el "
            f"archivo: {'; '.join(sospechosos[:5])}" + ("..." if len(sospechosos) > 5 else "")
        )
    return {
        "mensaje": mensaje,
        "total": len(filas),
        "sospechosos": sospechosos,
    }


@app.get("/productos-estrella")
def listar_productos_estrella(request: Request, campana: Optional[str] = None):
    query = "SELECT * FROM producto_estrella WHERE 1=1"
    params: list = []
    if campana:
        query += " AND campana = ?"
        params.append(campana)
    query += " ORDER BY competidor, descripcion_competidor"
    with get_conn() as conn:
        umbral = _umbral_alerta_actual(conn)
        rows = [dict(r) for r in conn.execute(query, params).fetchall()]
        resultado = []
        for r in rows:
            r["foto_url"] = (
                f"{request.base_url}static/productos_estrella_fotos/{r['foto_competidor']}"
                if r.get("foto_competidor") else None
            )
            r["foto_azzorti_url"] = (
                f"{request.base_url}static/productos_estrella_fotos/{r['foto_azzorti']}"
                if r.get("foto_azzorti") else None
            )
            r["campana_anterior"] = r.get("campana_azzorti")
            # El Excel trae varios decimales (ej. 4.05832032) - se redondea
            # a 2 decimales para mostrar, tanto los precios como el %.
            if r["precio_competidor"] is not None:
                r["precio_competidor"] = round(r["precio_competidor"], 2)
            if r["precio_azzorti"] is not None:
                r["precio_azzorti"] = round(r["precio_azzorti"], 2)
            # La columna se llama "Precio" en el archivo en los dos modos -
            # se respeta como precio y el % se calcula igual en los dos, en
            # vez de tratarlo distinto según modo.
            if r["precio_competidor"] is not None and r["precio_azzorti"]:
                delta = (r["precio_azzorti"] - r["precio_competidor"]) / r["precio_azzorti"] * 100
                r["delta_pct"] = round(delta, 2)
                r["precio_campana_anterior"] = None
                if r["modo"] == "HOMOLOGO_FIJO":
                    r["comparacion"] = "vs Azzorti (" + (r["azzorti_referente"] or "") + ")"
                else:
                    r["comparacion"] = "vs campaña anterior (" + (r["campana_azzorti"] or "") + ", " + (r["azzorti_referente"] or "") + ")"
            elif r["modo"] == "HOMOLOGO_FIJO":
                r["delta_pct"] = None
                r["precio_campana_anterior"] = None
                r["comparacion"] = "vs Azzorti (" + (r["azzorti_referente"] or "") + ") — falta precio"
            else:
                # Archivos viejos que no traen precio en esa columna: se
                # busca el precio de la otra campaña ya cargada para el
                # mismo producto, como respaldo.
                anterior = conn.execute(
                    """SELECT precio_competidor, campana FROM producto_estrella
                    WHERE competidor = ? AND descripcion_competidor = ? AND campana != ?
                    ORDER BY actualizado_en DESC LIMIT 1""",
                    (r["competidor"], r["descripcion_competidor"], r["campana"]),
                ).fetchone()
                r["precio_campana_anterior"] = round(anterior["precio_competidor"], 2) if anterior and anterior["precio_competidor"] is not None else None
                r["campana_anterior"] = anterior["campana"] if anterior else r["campana_anterior"]
                if anterior and anterior["precio_competidor"] and r["precio_competidor"] is not None:
                    delta = (r["precio_competidor"] - anterior["precio_competidor"]) / anterior["precio_competidor"] * 100
                    r["delta_pct"] = round(delta, 2)
                    r["comparacion"] = "vs campaña anterior (Bs " + str(anterior["precio_competidor"]) + ")"
                else:
                    r["delta_pct"] = None
                    r["comparacion"] = "vs campaña anterior — sin dato previo o precio actual pendiente"
            # Umbral unico (Pendiente 3): mismo criterio de alerta en los dos
            # modos, no una politica distinta por categoria/competidor.
            r["alerta"] = (r["delta_pct"] is not None) and (abs(r["delta_pct"]) > umbral)
            resultado.append(r)
    return resultado


@app.post("/capturas", status_code=201)
def crear_captura(c: CapturaIn):
    # Protege contra el timeout corto del cliente (8s en la app, ver
    # lib/main.dart): si la subida de la foto tarda mas que eso por el
    # tunel/red movil, la app le avisa al analista que "no se pudo
    # conectar" aunque el servidor si la termine guardando - el analista,
    # creyendo que fallo, repite la captura completa y eso duplica el
    # registro. Si llega una captura practicamente identica (mismo canal,
    # campaña, categoria, descripcion y precio) a los pocos minutos, se
    # devuelve la que ya existe en vez de crear una nueva.
    with get_conn() as conn:
        reciente = conn.execute(
            """SELECT id FROM captura
               WHERE competidor = ? AND canal = ? AND campana = ? AND categoria = ?
               AND descripcion = ? AND precio = ?
               AND creada_en > ?
               ORDER BY id DESC LIMIT 1""",
            (
                c.competidor, c.canal, c.campana, c.categoria,
                c.descripcion, c.precio,
                (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat(),
            ),
        ).fetchone()
        if reciente:
            return {
                "id": reciente["id"],
                "mensaje": "Ya se habia sincronizado este producto hace unos minutos - no se duplico.",
            }

    foto_archivo = None
    if c.foto_producto_b64:
        try:
            foto_archivo = _guardar_foto_captura(c.competidor, c.foto_producto_b64)
        except Exception:
            foto_archivo = None  # best-effort: si la foto falla, la captura sigue

    with get_conn() as conn:
        try:
            cur = conn.execute(
                """INSERT INTO captura
                (competidor, canal, campana, categoria, nivel_precio, descripcion,
                 silueta, talla, composicion1, composicion2, manga, color, detalle,
                 caracteristicas, precio, sku_competidor, foto_archivo, creada_en, catalogo_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    c.competidor, c.canal, c.campana, c.categoria, c.nivel_precio,
                    c.descripcion, c.silueta, c.talla, c.composicion1, c.composicion2,
                    c.manga, c.color, c.detalle, c.caracteristicas, c.precio,
                    c.sku_competidor, foto_archivo,
                    datetime.now(timezone.utc).isoformat(), c.catalogo_id,
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(
                status_code=409,
                detail=f"Registro duplicado: ya existe una captura de '{c.competidor}' "
                       f"con SKU '{c.sku_competidor}' en la campaña '{c.campana}'.",
            )
        captura_id = cur.lastrowid
    return {"id": captura_id, "mensaje": "Captura sincronizada correctamente."}


@app.get("/capturas")
def listar_capturas(
    request: Request, competidor: Optional[str] = None, campana: Optional[str] = None
):
    query = "SELECT * FROM captura WHERE 1=1"
    params: list = []
    if competidor:
        query += " AND competidor = ?"
        params.append(competidor)
    if campana:
        query += " AND campana = ?"
        params.append(campana)
    query += " ORDER BY id DESC"
    with get_conn() as conn:
        umbral = _umbral_alerta_actual(conn)
        rows = conn.execute(query, params).fetchall()
        resultado = []
        for r in rows:
            d = dict(r)
            d["foto_url"] = (
                f"{request.base_url}static/capturas_fotos/{d['foto_archivo']}"
                if d.get("foto_archivo") else None
            )
            # Si el producto estrella se registró a partir de un catálogo
            # subido (Venta Directa), se adjunta el enlace para volver a
            # verlo — el analista necesita poder abrir la página original.
            if d.get("catalogo_id"):
                catalogo = conn.execute(
                    "SELECT * FROM catalogo_competidor WHERE id = ?",
                    (d["catalogo_id"],),
                ).fetchone()
                if catalogo:
                    d["catalogo"] = {
                        "id": catalogo["id"],
                        "competidor": catalogo["competidor"],
                        "campana": catalogo["campana"],
                        "archivo_url": f"{request.base_url}static/catalogos_competidor/{catalogo['archivo']}",
                    }
            # Si ya se confirmó homologación, se adjunta el resultado de la
            # comparación de precio — umbral unico (Pendiente 3), ya no hay
            # politica puntual por categoria/competidor.
            if d.get("azzorti_sku_confirmado") and d["canal"] == "Venta Directa":
                # Venta Directa homologa contra el catalogo indexado (ver
                # _sugerir_homologacion_venta_directa), no contra
                # azzorti_producto (eso es solo Retail) - antes esta tabla
                # se quedaba buscando en el lugar equivocado y la
                # homologacion confirmada nunca se reflejaba aqui, aunque
                # si estuviera guardada en la base de datos.
                prod = conn.execute(
                    "SELECT * FROM catalogo_producto WHERE producto_codigo = ? "
                    "AND precio IS NOT NULL ORDER BY id DESC LIMIT 1",
                    (d["azzorti_sku_confirmado"],),
                ).fetchone()
                if prod:
                    delta = (prod["precio"] - d["precio"]) / prod["precio"] * 100
                    catalogo_prod = conn.execute(
                        "SELECT id FROM catalogo_competidor WHERE LOWER(TRIM(competidor)) = 'azzorti' "
                        "ORDER BY id DESC LIMIT 1"
                    ).fetchone()
                    foto_pagina = CATALOGO_PAGINAS_DIR / f"{catalogo_prod['id']}_{prod['pagina']}_{prod['producto_codigo']}.png" if catalogo_prod else None
                    d["homologacion"] = {
                        "azzorti_sku": prod["producto_codigo"],
                        "azzorti_descripcion": (prod["texto_cercano"] or "")[:80] or prod["producto_codigo"],
                        "pagina_catalogo": prod["pagina"],
                        "precio_azzorti": prod["precio"],
                        "delta_pct": round(delta, 1),
                        "alerta_generica": abs(delta) > umbral,
                        "foto_url": (
                            f"{request.base_url}static/catalogo_paginas/{foto_pagina.name}"
                            if foto_pagina and foto_pagina.exists() else None
                        ),
                    }
            elif d.get("azzorti_sku_confirmado"):
                azzorti = conn.execute(
                    "SELECT * FROM azzorti_producto WHERE sku = ?",
                    (d["azzorti_sku_confirmado"],),
                ).fetchone()
                if azzorti:
                    delta = (azzorti["precio"] - d["precio"]) / azzorti["precio"] * 100
                    d["homologacion"] = {
                        "azzorti_sku": azzorti["sku"],
                        "azzorti_descripcion": azzorti["descripcion"],
                        "pagina_catalogo": azzorti["pagina_catalogo"],
                        "precio_azzorti": azzorti["precio"],
                        "delta_pct": round(delta, 1),
                        "alerta_generica": abs(delta) > umbral,
                        "foto_url": (
                            f"{request.base_url}static/{azzorti['foto_archivo']}"
                            if azzorti["foto_archivo"] else None
                        ),
                    }
                else:
                    # Retail tambien puede homologar contra un producto
                    # indexado del catalogo PDF real de Azzorti (mismo
                    # catalogo que usa Venta Directa, seccion Moda - ver
                    # _candidatos_moda_azzorti), no solo contra los 19
                    # productos de muestra de azzorti_producto.
                    prod = conn.execute(
                        "SELECT * FROM catalogo_producto WHERE producto_codigo = ? "
                        "AND precio IS NOT NULL ORDER BY id DESC LIMIT 1",
                        (d["azzorti_sku_confirmado"],),
                    ).fetchone()
                    if prod:
                        delta = (prod["precio"] - d["precio"]) / prod["precio"] * 100
                        catalogo_azz = conn.execute(
                            "SELECT id FROM catalogo_competidor WHERE LOWER(TRIM(competidor)) = 'azzorti' "
                            "ORDER BY id DESC LIMIT 1"
                        ).fetchone()
                        foto_prod = (
                            CATALOGO_PAGINAS_DIR / f"{catalogo_azz['id']}_{prod['pagina']}_{prod['producto_codigo']}.png"
                            if catalogo_azz else None
                        )
                        d["homologacion"] = {
                            "azzorti_sku": prod["producto_codigo"],
                            "azzorti_descripcion": (prod["texto_cercano"] or "")[:80] or prod["producto_codigo"],
                            "pagina_catalogo": prod["pagina"],
                            "precio_azzorti": prod["precio"],
                            "delta_pct": round(delta, 1),
                            "alerta_generica": abs(delta) > umbral,
                            "foto_url": (
                                f"{request.base_url}static/catalogo_paginas/{foto_prod.name}"
                                if foto_prod and foto_prod.exists() else None
                            ),
                        }
            elif d.get("descripcion"):
                # Sin homologacion: Pendiente 3 pide comparar contra la
                # misma captura (mismo competidor + descripcion) de la
                # campana anterior — antes esto solo existia para Venta
                # Directa (producto_estrella); ahora aplica igual en Retail.
                anterior = conn.execute(
                    """SELECT precio, campana FROM captura
                    WHERE competidor = ? AND descripcion = ? AND campana != ? AND id != ?
                    ORDER BY creada_en DESC LIMIT 1""",
                    (d["competidor"], d["descripcion"], d["campana"], d["id"]),
                ).fetchone()
                if anterior:
                    delta = (d["precio"] - anterior["precio"]) / anterior["precio"] * 100
                    d["vs_campana_anterior"] = {
                        "campana_anterior": anterior["campana"],
                        "precio_anterior": anterior["precio"],
                        "delta_pct": round(delta, 1),
                        "alerta_generica": abs(delta) > umbral,
                    }
            resultado.append(d)
    return resultado


# Familias de silueta: "Entallada" y "Semiajustada" describen básicamente
# la misma idea (prenda ceñida al cuerpo) con palabras distintas — antes se
# exigía coincidencia exacta de texto, así que nunca se cruzaban aunque el
# corte fuera prácticamente el mismo.
_FAMILIAS_SILUETA = [
    {"entallada", "semiajustada", "ajustada", "ceñida"},
    {"suelta", "oversize", "holgada", "amplia"},
    {"recta"},
]

# Palabras que no aportan a comparar el tipo de tela (números de
# porcentaje, conectores) — se ignoran para que la comparación de
# composición se enfoque en qué tela es, no en el % exacto de cada una
# (que casi nunca va a coincidir entre competencia y Azzorti de todas
# formas, ya que son prendas distintas).
def _tokens_tela(texto: str) -> set:
    palabras = texto.lower().replace("+", " ").split()
    return {p.strip("%.,") for p in palabras if not p.rstrip("%").replace(".", "").isdigit()}


def _familia_silueta(valor: str) -> set | None:
    for familia in _FAMILIAS_SILUETA:
        if valor in familia:
            return familia
    return None


def _score_similitud(captura: sqlite3.Row, producto: sqlite3.Row) -> float:
    """Puntaje 0-100. Categoría ya viene filtrada como condición obligatoria;
    esto puntúa qué tan cerca están los atributos que sí capturó el analista.
    Nunca compara por código — solo por atributos y color, tal como se
    definió con Yohana (no hay SKU compartido con la competencia)."""
    score = 0.0
    color_c = (captura["color"] or "").strip().lower()
    color_p = (producto["color"] or "").strip().lower()
    if color_c and color_p and color_c == color_p:
        score += 30

    silueta_c = (captura["silueta"] or "").strip().lower()
    silueta_p = (producto["silueta"] or "").strip().lower()
    if silueta_c and silueta_p:
        if silueta_c == silueta_p:
            score += 30
        else:
            familia_c = _familia_silueta(silueta_c)
            if familia_c and silueta_p in familia_c:
                score += 18  # mismo concepto de corte, distinta palabra

    comp_c = f"{captura['composicion1'] or ''} {captura['composicion2'] or ''}"
    comp_p = producto["composicion"] or ""
    if comp_c.strip() and comp_p.strip():
        tokens_c = _tokens_tela(comp_c)
        tokens_p = _tokens_tela(comp_p)
        union = tokens_c | tokens_p
        interseccion = tokens_c & tokens_p
        if interseccion and union:
            # Índice de Jaccard: qué proporción de todas las telas
            # mencionadas (por cualquiera de los dos lados) coincide —
            # más justo que dividir solo por el lado de Azzorti, que
            # castigaba de más cuando el competidor mencionaba más telas.
            score += 25 * len(interseccion) / len(union)

    manga_c = (captura["manga"] or "").strip().lower()
    manga_p = (producto["manga"] or "").strip().lower()
    if manga_c and manga_p and manga_c == manga_p:
        score += 15

    return round(score, 1)


# El mismo catalogo PDF de Azzorti (206 paginas) sirve tanto para Venta
# Directa como para Retail (confirmado por Yohana 2026-08-11) - la
# seccion "MODA" trae prendas reales que complementan los 19 productos
# de muestra de azzorti_producto. No hay categoria estructurada por fila
# en catalogo_producto (a diferencia de azzorti_producto), asi que se
# exige que una palabra clave del tipo de prenda aparezca en el texto
# OCR cercano al codigo - sin esto, un jean y una blusa competirian por
# igual solo por compartir "algodon" en la composicion.
_CATEGORIA_RETAIL_PALABRA_CLAVE = {
    "blusas femeninas": "BLUSA",
    "vestidos femeninos cortos": "VESTIDO",
    "vestidos femeninos largos": "VESTIDO",
    "camisetas masculinas": "CAMISETA",
    "polos masculinos": "POLO",
    "camisetas femeninas": "CAMISETA",
    "crop top femenino": "CROP",
    "jeans femeninos": "JEAN",
    "jeans masculinos": "JEAN",
    "lenceria ppp": "LENCERIA",
    "cubrecamas dobles": "CUBRECAMA",
    "sabanas dobles": "SABANA",
    "toallas": "TOALLA",
    "mochilas infantiles": "MOCHILA",
    "cubrecamas sencillos infantiles": "CUBRECAMA",
}


def _candidatos_moda_azzorti(conn, categoria_captura: str) -> list[sqlite3.Row]:
    palabra = _CATEGORIA_RETAIL_PALABRA_CLAVE.get(_sin_acentos((categoria_captura or "").strip().lower()))
    if not palabra:
        return []
    filas = conn.execute(
        """SELECT cp.* FROM catalogo_producto cp
        JOIN catalogo_competidor cc ON cc.id = cp.catalogo_id
        WHERE LOWER(TRIM(cc.competidor)) = 'azzorti'
        AND (cp.seccion IS NULL OR UPPER(cp.seccion) LIKE '%MODA%')"""
    ).fetchall()
    return [f for f in filas if palabra in _sin_acentos((f["texto_cercano"] or "").upper())]


def _sugerir_homologacion_venta_directa(conn, captura: sqlite3.Row, request: Request) -> dict:
    """Homologa una captura de Venta Directa contra el catalogo real de
    Azzorti subido en Competencia (competidor='Azzorti') - no hay ningun
    Excel estructurado de productos Azzorti de Venta Directa (a
    diferencia de Retail, que si tiene 'azzorti_producto'), asi que se
    usa el catalogo PDF indexado por OCR (ver POST /catalogos/{id}/
    indexar-productos) y se compara por palabras clave contra el texto
    que quedo cerca de cada codigo de producto."""
    catalogo = conn.execute(
        "SELECT * FROM catalogo_competidor WHERE LOWER(TRIM(competidor)) = 'azzorti' "
        "AND campana = ? ORDER BY id DESC LIMIT 1",
        (captura["campana"],),
    ).fetchone()
    if not catalogo:
        catalogo = conn.execute(
            "SELECT * FROM catalogo_competidor WHERE LOWER(TRIM(competidor)) = 'azzorti' "
            "ORDER BY id DESC LIMIT 1"
        ).fetchone()
    if not catalogo:
        return {
            "captura_id": captura["id"],
            "criterio": "No hay ningún catálogo de Azzorti subido en Competencia todavía.",
            "sugerencias": [],
        }
    todos = conn.execute(
        "SELECT * FROM catalogo_producto WHERE catalogo_id = ?", (catalogo["id"],)
    ).fetchall()
    if not todos:
        return {
            "captura_id": captura["id"],
            "criterio": f"El catálogo de Azzorti '{catalogo['archivo']}' todavía no fue indexado "
                        "(POST /catalogos/{id}/indexar-productos) o no se encontró ningún producto en él.",
            "sugerencias": [],
        }
    # Filtra por seccion del catalogo antes de comparar texto - sin esto,
    # una captura de Fragancias terminaba comparandose contra sabanas,
    # toallas y hasta texto de ingredientes de otra pagina, solo porque
    # coincidia un numero suelto en el texto OCR.
    candidatos = [p for p in todos if _candidato_coincide_categoria(captura["categoria"], p["seccion"])]
    texto_principal = " ".join((captura["categoria"] or "", captura["descripcion"] or ""))
    texto_secundario = " ".join((captura["caracteristicas"] or "", captura["detalle"] or ""))

    def _foto_url_producto(pagina: int, codigo: str) -> Optional[str]:
        ruta_foto = CATALOGO_PAGINAS_DIR / f"{catalogo['id']}_{pagina}_{codigo}.png"
        return f"{request.base_url}static/catalogo_paginas/{ruta_foto.name}" if ruta_foto.exists() else None

    sugerencias = sorted(
        (
            {
                "sku": p["producto_codigo"],
                "categoria": captura["categoria"],
                "descripcion": p["texto_cercano"][:80] if p["texto_cercano"] else p["producto_codigo"],
                "color": None,
                "composicion": None,
                "silueta": None,
                "manga": None,
                "precio": p["precio"] or 0,
                "pagina_catalogo": p["pagina"],
                "foto_url": _foto_url_producto(p["pagina"], p["producto_codigo"]),
                "score_similitud": round(_score_texto(texto_principal, texto_secundario, p["texto_cercano"] or "") * 100, 1),
            }
            for p in candidatos
        ),
        key=lambda x: x["score_similitud"],
        reverse=True,
    )
    sugerencias = [s for s in sugerencias if s["score_similitud"] > 0]
    # Dos codigos de la misma pagina comparten la misma foto (es una foto
    # por pagina, no por producto) y a veces hasta el texto cercano se
    # mezcla entre ellos si estan pegados en el catalogo - mostrar los dos
    # como tarjetas separadas se ve como resultados duplicados/rotos. Se
    # deja solo el de mejor puntaje por pagina.
    vistas = set()
    sugerencias_por_pagina = []
    for s in sugerencias:
        if s["pagina_catalogo"] in vistas:
            continue
        vistas.add(s["pagina_catalogo"])
        sugerencias_por_pagina.append(s)
    sugerencias = sugerencias_por_pagina[:20]
    return {
        "captura_id": captura["id"],
        "criterio": f"Comparado por palabras clave contra el catálogo de Azzorti "
                    f"'{catalogo['archivo']}' (campaña {catalogo['campana']}) — no hay ficha de "
                    "atributos por producto en Venta Directa como sí la hay en Retail, así que la "
                    "coincidencia es aproximada (texto OCR cercano al código de cada producto).",
        "sugerencias": sugerencias,
    }


@app.get("/capturas/{captura_id}/homologacion/sugerencias")
def sugerir_homologacion(captura_id: int, request: Request):
    with get_conn() as conn:
        captura = conn.execute("SELECT * FROM captura WHERE id = ?", (captura_id,)).fetchone()
        if not captura:
            raise HTTPException(404, "Captura no encontrada")

        if captura["canal"] == "Venta Directa":
            return _sugerir_homologacion_venta_directa(conn, captura, request)

        # Filtra solo por categoria: el catalogo real de Azzorti no tiene
        # niveles Bajo/Medio/Alto (eso era propio de como se agrupaba a la
        # competencia). El nivel_precio de la captura queda como dato
        # descriptivo de lo que vio el analista, no como llave de cruce.
        candidatos = conn.execute(
            "SELECT * FROM azzorti_producto WHERE categoria = ?",
            (captura["categoria"],),
        ).fetchall()
        # Complementa los 19 productos de muestra con productos reales
        # indexados del mismo catalogo PDF de Azzorti que usa Venta Directa
        # (confirmado por Yohana: es el mismo catalogo para los dos
        # canales) - sin esto, categorias con pocos o ningun producto de
        # muestra (ej. solo 3 "Blusas Femeninas") no tenian con que
        # homologar de verdad.
        skus_muestra = {p["sku"] for p in candidatos}
        candidatos_moda = [
            p for p in _candidatos_moda_azzorti(conn, captura["categoria"])
            if p["producto_codigo"] not in skus_muestra
        ]
        catalogo_azzorti = conn.execute(
            "SELECT id FROM catalogo_competidor WHERE LOWER(TRIM(competidor)) = 'azzorti' ORDER BY id DESC LIMIT 1"
        ).fetchone()

    def _foto_url_moda(p) -> Optional[str]:
        if not catalogo_azzorti:
            return None
        ruta = CATALOGO_PAGINAS_DIR / f"{catalogo_azzorti['id']}_{p['pagina']}_{p['producto_codigo']}.png"
        return f"{request.base_url}static/catalogo_paginas/{ruta.name}" if ruta.exists() else None

    sugerencias_muestra = [
        {
            "sku": p["sku"],
            "categoria": p["categoria"],
            "descripcion": p["descripcion"],
            "color": p["color"],
            "composicion": p["composicion"],
            "silueta": p["silueta"],
            "manga": p["manga"],
            "precio": p["precio"],
            "pagina_catalogo": p["pagina_catalogo"],
            "foto_url": (
                f"{request.base_url}static/{p['foto_archivo']}"
                if p["foto_archivo"] else None
            ),
            "score_similitud": _score_similitud(captura, p),
        }
        for p in candidatos
    ]
    sugerencias_moda = [
        {
            "sku": p["producto_codigo"],
            "categoria": captura["categoria"],
            "descripcion": (p["texto_cercano"] or "")[:80] or p["producto_codigo"],
            "color": None,
            "composicion": p["texto_cercano"],
            "silueta": None,
            "manga": None,
            "precio": p["precio"] or 0,
            "pagina_catalogo": p["pagina"],
            "foto_url": _foto_url_moda(p),
            "score_similitud": _score_similitud(
                captura, {"color": None, "silueta": None, "composicion": p["texto_cercano"], "manga": None}
            ),
        }
        for p in candidatos_moda
    ]
    sugerencias = sorted(sugerencias_muestra + sugerencias_moda, key=lambda x: x["score_similitud"], reverse=True)
    return {
        "captura_id": captura_id,
        "criterio": "Filtrado por categoría, combinando el catálogo de muestra "
                    "(azzorti_producto) con productos reales indexados del catálogo PDF "
                    "de Azzorti (mismo catálogo que usa Venta Directa). Ranking por "
                    "color, silueta, composición y manga — nunca por código, ya que "
                    "la competencia no comparte SKU con Azzorti. Los productos indexados "
                    "del PDF no tienen color/silueta/manga estructurados, así que solo "
                    "puntúan por composición (tela).",
        "sugerencias": sugerencias,
    }


@app.post("/capturas/{captura_id}/homologacion/confirmar")
def confirmar_homologacion(captura_id: int, body: HomologacionConfirmar):
    with get_conn() as conn:
        captura = conn.execute("SELECT * FROM captura WHERE id = ?", (captura_id,)).fetchone()
        if not captura:
            raise HTTPException(404, "Captura no encontrada")
        if captura["canal"] == "Venta Directa":
            # En Venta Directa "azzorti_sku_confirmado" guarda el codigo de
            # producto del catalogo_producto indexado, no un SKU de
            # azzorti_producto (Retail) - mismo campo, distinta fuente.
            existe = conn.execute(
                "SELECT 1 FROM catalogo_producto WHERE producto_codigo = ?", (body.azzorti_sku,)
            ).fetchone()
            if not existe:
                raise HTTPException(404, f"Código '{body.azzorti_sku}' no existe en el catálogo de Azzorti indexado")
        else:
            # Retail ahora tambien puede homologar contra un producto
            # indexado del catalogo PDF (ver _candidatos_moda_azzorti), no
            # solo contra los 19 de muestra en azzorti_producto.
            existe = conn.execute(
                "SELECT 1 FROM azzorti_producto WHERE sku = ? "
                "UNION SELECT 1 FROM catalogo_producto WHERE producto_codigo = ?",
                (body.azzorti_sku, body.azzorti_sku),
            ).fetchone()
            if not existe:
                raise HTTPException(404, f"SKU Azzorti '{body.azzorti_sku}' no existe en el catálogo")
        conn.execute(
            "UPDATE captura SET azzorti_sku_confirmado = ? WHERE id = ?",
            (body.azzorti_sku, captura_id),
        )
    return {"mensaje": f"Homologación confirmada: captura {captura_id} <-> {body.azzorti_sku}"}


@app.get("/capturas/{captura_id}/evaluacion")
def evaluar_captura(captura_id: int):
    """Evalua una captura contra Azzorti (si ya tiene homologacion) o, si no
    tiene, contra su propio precio de la campana anterior (Pendiente 3: esta
    regla ahora aplica igual en Retail que en Venta Directa). Ya no consulta
    politica_precio - un solo umbral editable decide la alerta en los dos
    casos, sin distinguir categoria/competidor."""
    with get_conn() as conn:
        captura = conn.execute("SELECT * FROM captura WHERE id = ?", (captura_id,)).fetchone()
        if not captura:
            raise HTTPException(404, "Captura no encontrada")
        umbral = _umbral_alerta_actual(conn)
        precio_competencia = captura["precio"]

        if captura["azzorti_sku_confirmado"]:
            if captura["canal"] == "Venta Directa":
                azzorti = conn.execute(
                    "SELECT producto_codigo AS sku, precio FROM catalogo_producto WHERE producto_codigo = ? "
                    "AND precio IS NOT NULL ORDER BY id DESC LIMIT 1",
                    (captura["azzorti_sku_confirmado"],),
                ).fetchone()
                if not azzorti:
                    raise HTTPException(
                        404,
                        "El código Azzorti confirmado ya no existe en el catálogo indexado, o no "
                        "se le pudo leer un precio por OCR.",
                    )
            else:
                azzorti = conn.execute(
                    "SELECT * FROM azzorti_producto WHERE sku = ?", (captura["azzorti_sku_confirmado"],)
                ).fetchone()
                if not azzorti:
                    # Retail tambien puede homologar contra un producto
                    # indexado del catalogo PDF (ver _candidatos_moda_azzorti).
                    prod = conn.execute(
                        "SELECT producto_codigo AS sku, precio FROM catalogo_producto "
                        "WHERE producto_codigo = ? AND precio IS NOT NULL ORDER BY id DESC LIMIT 1",
                        (captura["azzorti_sku_confirmado"],),
                    ).fetchone()
                    if not prod:
                        raise HTTPException(404, "El SKU Azzorti confirmado ya no existe en el catálogo.")
                    azzorti = prod
            delta_pct = (azzorti["precio"] - precio_competencia) / azzorti["precio"] * 100
            return {
                "captura_id": captura_id,
                "modo": "HOMOLOGO",
                "azzorti_sku": azzorti["sku"],
                "precio_azzorti": azzorti["precio"],
                "precio_competencia": precio_competencia,
                "delta_pct": round(delta_pct, 1),
                "umbral_pct": umbral,
                "alerta": abs(delta_pct) > umbral,
            }

        anterior = conn.execute(
            """SELECT precio, campana FROM captura
            WHERE competidor = ? AND descripcion = ? AND campana != ? AND id != ?
            ORDER BY creada_en DESC LIMIT 1""",
            (captura["competidor"], captura["descripcion"], captura["campana"], captura_id),
        ).fetchone()
        if not anterior:
            return {
                "captura_id": captura_id,
                "modo": "SIN_DATO",
                "precio_competencia": precio_competencia,
                "umbral_pct": umbral,
                "mensaje": "Sin homologación y sin captura previa del mismo producto para comparar.",
            }
        delta_pct = (precio_competencia - anterior["precio"]) / anterior["precio"] * 100
        return {
            "captura_id": captura_id,
            "modo": "VS_CAMPANA_ANTERIOR",
            "campana_anterior": anterior["campana"],
            "precio_anterior": anterior["precio"],
            "precio_competencia": precio_competencia,
            "delta_pct": round(delta_pct, 1),
            "umbral_pct": umbral,
            "alerta": abs(delta_pct) > umbral,
        }


def _anio_de_campana(campana: str) -> int:
    """Saca el año de una campaña de Venta Directa (ej. 'C09 2026' -> 2026).
    Retail todavia no guarda el año como parte de la campaña (ej. 'C-10
    (activa)' no dice de que año es) - mientras eso no se defina, se asume
    el año en curso. Es una aproximacion a proposito visible/documentada,
    no un dato inventado en silencio."""
    m = re.search(r"(20\d{2})", campana or "")
    if m:
        return int(m.group(1))
    return datetime.now(timezone.utc).year


@app.get("/historico")
def historico():
    """Evolución de la Variación % vs Azzorti por competidor+campaña+canal,
    para la vista 'Precios'. Reemplaza la version anterior (que solo
    promediaba el precio propio del competidor, sin relacion con Azzorti -
    Yohana no le encontraba el fin porque no decia nada sobre la brecha con
    Azzorti). Solo se promedian productos que SI tienen homologacion
    confirmada contra Azzorti (no los 'vs campaña anterior', que comparan
    contra el propio precio del competidor, no contra Azzorti). Tambien
    devuelve cuantos de esos productos dispararon alerta de umbral, para el
    panel de alertas historicas."""
    deltas = []
    with get_conn() as conn:
        umbral = _umbral_alerta_actual(conn)
        for c in conn.execute("SELECT * FROM captura WHERE azzorti_sku_confirmado IS NOT NULL"):
            precio_azzorti = None
            if c["canal"] == "Venta Directa":
                prod = conn.execute(
                    "SELECT precio FROM catalogo_producto WHERE producto_codigo = ? "
                    "AND precio IS NOT NULL ORDER BY id DESC LIMIT 1",
                    (c["azzorti_sku_confirmado"],),
                ).fetchone()
                precio_azzorti = prod["precio"] if prod else None
            else:
                azz = conn.execute(
                    "SELECT precio FROM azzorti_producto WHERE sku = ?", (c["azzorti_sku_confirmado"],)
                ).fetchone()
                if azz:
                    precio_azzorti = azz["precio"]
                else:
                    prod = conn.execute(
                        "SELECT precio FROM catalogo_producto WHERE producto_codigo = ? "
                        "AND precio IS NOT NULL ORDER BY id DESC LIMIT 1",
                        (c["azzorti_sku_confirmado"],),
                    ).fetchone()
                    precio_azzorti = prod["precio"] if prod else None
            if precio_azzorti and c["precio"] is not None:
                delta = (precio_azzorti - c["precio"]) / precio_azzorti * 100
                deltas.append({
                    "canal": "Retail" if c["canal"] != "Venta Directa" else "VentaDirecta",
                    "competidor": c["competidor"], "campana": c["campana"],
                    "delta_pct": delta, "alerta": abs(delta) > umbral,
                })
        for p in conn.execute(
            "SELECT * FROM producto_estrella WHERE modo = 'HOMOLOGO_FIJO' "
            "AND precio_azzorti IS NOT NULL AND precio_competidor IS NOT NULL"
        ):
            delta = (p["precio_azzorti"] - p["precio_competidor"]) / p["precio_azzorti"] * 100
            deltas.append({
                "canal": "VentaDirecta", "competidor": p["competidor"], "campana": p["campana"],
                "delta_pct": delta, "alerta": abs(delta) > umbral,
            })

    grupos: dict = {}
    for d in deltas:
        key = (d["canal"], d["competidor"], d["campana"])
        grupos.setdefault(key, []).append(d)
    filas = []
    for (canal, competidor, campana), items in grupos.items():
        alertas = [i for i in items if i["alerta"]]
        filas.append({
            "canal": canal,
            "competidor": competidor,
            "campana": campana,
            "anio": _anio_de_campana(campana),
            "delta_pct_promedio": round(sum(i["delta_pct"] for i in items) / len(items), 1),
            "cantidad_productos": len(items),
            "cantidad_alertas": len(alertas),
        })
    return filas


# ====================== PENDIENTE 8: IDENTIFICACION DE OFERTAS ======================
#
# El catalogo real de un competidor (PDF) es una sola imagen plana por
# pagina - foto de producto, banda de oferta y texto quedan mezclados en
# un solo grafico, no hay un logotipo aislado que se pueda comparar por
# similitud de pixeles contra las imagenes de "Detalle de ofertas.xlsx".
# Ademas los nombres de oferta que Mercadeo registro ahi ("OFERTA WOW")
# no coinciden letra por letra con el texto real de la banda en el
# catalogo ("WOW FEST", "WOW!") - son la misma familia de oferta pero no
# el mismo string. Por eso la deteccion es: OCR de cada pagina + coincidencia
# por palabras clave contra el nombre de oferta de referencia (filtrado por
# competidor), no comparacion de imagen. Acordado con Yohana 2026-08-06.

_STOPWORDS_OFERTA = {
    "MAS", "DEL", "DE", "LA", "EL", "LOS", "LAS", "UN", "UNA", "Y", "O",
    "A", "AL", "OFERTA", "OFERTAS", "EN",
    # Conectores genericos de cualquier descripcion de producto (ej. "con
    # protector solar") - sin esto, una palabra tan comun como "CON"
    # empataba por igual contra productos sin ninguna relacion real,
    # mismo problema ya encontrado con "FEMENINA" en Pendiente de
    # homologacion Venta Directa.
    "CON", "SIN", "POR", "PARA", "QUE", "SE", "ES", "SU", "SUS", "MUY",
    "ESTE", "ESTA", "ESTOS", "ESTAS", "SUS", "LE", "LES", "COMO",
}


def _sin_acentos(texto: str) -> str:
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


_UNIDADES_TAMANO = {"ML", "MI", "OZ", "GR", "G", "CM", "KG", "FL", "LB", "MG"}


def _tokens_significativos(texto: str) -> set[str]:
    """Palabras/numeros relevantes de un texto para comparar ofertas -
    sin acentos, sin stopwords genericas ('oferta', 'del', ...) que
    apareceria en casi cualquier banda y no ayudan a distinguir una
    oferta de otra. Un numero seguido de una unidad ('50 ml') se descarta
    porque es el contenido del producto, no un porcentaje de descuento."""
    limpio = _sin_acentos((texto or "").upper())
    limpio = "".join(c if c.isalnum() else " " for c in limpio)
    palabras = limpio.split()
    tokens = set()
    for i, palabra in enumerate(palabras):
        if len(palabra) < 2 or palabra in _STOPWORDS_OFERTA:
            continue
        if palabra.isdigit() and i + 1 < len(palabras) and palabras[i + 1] in _UNIDADES_TAMANO:
            continue
        tokens.add(palabra)
    return tokens


def _coincide_token(ocr_tokens: set[str], ref_token: str) -> bool:
    """El OCR frecuentemente lee el signo '%' de una banda de oferta como
    un digito pegado (ej. 'con 507 desde' en vez de 'con 50% desde') - un
    numero de referencia tambien hace match si algun token OCR EMPIEZA con
    ese numero y tiene a lo sumo un caracter mal leido de mas (tope
    ajustado para no confundirlo con un precio o codigo mas largo)."""
    if ref_token in ocr_tokens:
        return True
    if ref_token.isdigit():
        return any(
            t.isdigit() and t.startswith(ref_token) and len(t) <= len(ref_token) + 1
            for t in ocr_tokens
        )
    return False


def _score_oferta(texto_ocr: str, nombre_oferta: str) -> float:
    """Fraccion de las palabras clave del nombre de oferta de referencia
    que aparecen en el texto que el OCR leyo de la pagina."""
    ref_tokens = _tokens_significativos(nombre_oferta)
    if not ref_tokens:
        return 0.0
    ocr_tokens = _tokens_significativos(texto_ocr)
    encontrados = sum(1 for t in ref_tokens if _coincide_token(ocr_tokens, t))
    return encontrados / len(ref_tokens)


UMBRAL_SCORE_OFERTA = 0.5  # al menos la mitad de las palabras clave deben coincidir


def _mejor_oferta(texto_ocr: str, ofertas_ref: list[sqlite3.Row]) -> tuple[Optional[str], float]:
    """Elige, entre las ofertas de referencia del mismo competidor, la que
    mejor coincide con el texto OCR de una pagina. Si ninguna llega al
    umbral, devuelve None en vez de forzar una coincidencia dudosa."""
    mejor_nombre, mejor_score = None, 0.0
    for ref in ofertas_ref:
        score = _score_oferta(texto_ocr, ref["nombre_oferta"])
        if score > mejor_score:
            mejor_nombre, mejor_score = ref["nombre_oferta"], score
    if mejor_score >= UMBRAL_SCORE_OFERTA:
        return mejor_nombre, mejor_score
    return None, mejor_score


# --- Pendiente 8 (v2): asignar la oferta al producto de la pagina, no a
# toda la pagina. Una pagina de catalogo puede traer mas de un producto
# (ej. una mochila y una lonchera en esquinas distintas) y la banda de
# oferta suele aplicar solo a uno de ellos - se usa la POSICION de cada
# palabra que el OCR reconocio (no solo el texto) para saber que "cod.
# NNNNN" esta mas cerca de la banda detectada, en vez de asumir que la
# oferta de la pagina aplica a todos los productos que aparecen en ella.

def _limpiar_palabra(palabra: str) -> str:
    limpio = _sin_acentos((palabra or "").upper())
    return "".join(c for c in limpio if c.isalnum())


_CODIGO_RE = re.compile(r"[A-Z]?\d{3,7}")


def _anclas_producto(data: dict) -> list[dict]:
    """Busca 'cod. NNNNN' palabra por palabra en la salida de
    pytesseract.image_to_data y devuelve la posicion (centro en pixeles)
    de cada codigo de producto que aparece en la pagina. El codigo puede
    ser puramente numerico (Venta Directa, ej. '392625') o con una letra
    de prefijo (Retail, ej. 'R2159') - antes solo se aceptaba el primero,
    asi que ningun producto de Moda quedaba indexado."""
    palabras = [_limpiar_palabra(t) for t in data["text"]]
    anclas = []
    for i, palabra in enumerate(palabras):
        if not palabra.startswith("COD") or len(palabra) > 5:
            continue
        for j in range(i + 1, min(i + 3, len(palabras))):
            if _CODIGO_RE.fullmatch(palabras[j]):
                anclas.append({
                    "codigo": palabras[j],
                    "x": data["left"][j] + data["width"][j] / 2,
                    "y": data["top"][j] + data["height"][j] / 2,
                })
                break
    return anclas


_REF_RE = re.compile(r"REF([A-Z]?\d{3,7})")


def _anclas_producto_retail(data: dict) -> list[dict]:
    """Las paginas de Moda/Retail del mismo catalogo NO usan 'COD.
    NNNNN' como Venta Directa - usan 'Nombre Ref.RNNNN' junto al precio,
    con una tabla de codigos por talla aparte (esos codigos de talla se
    ignoran a proposito, no son el SKU que usa azzorti_producto). El OCR
    junta 'Ref.R1411' en un solo token al limpiar el punto, asi que se
    busca el patron REF+codigo pegado en la misma palabra."""
    anclas = []
    for i, t in enumerate(data["text"]):
        m = _REF_RE.fullmatch(_limpiar_palabra(t))
        if m:
            anclas.append({
                "codigo": m.group(1),
                "x": data["left"][i] + data["width"][i] / 2,
                "y": data["top"][i] + data["height"][i] / 2,
            })
    return anclas


def _ocurrencias_token(data: dict, token: str) -> list[dict]:
    """Posicion de cada palabra OCR de la pagina que coincide con una
    palabra clave de una oferta de referencia (mismo criterio que
    _coincide_token, pero palabra por palabra para saber DONDE aparecio)."""
    ocurrencias = []
    for i, texto in enumerate(data["text"]):
        palabra = _limpiar_palabra(texto)
        if palabra and _coincide_token({palabra}, token):
            ocurrencias.append({
                "x": data["left"][i] + data["width"][i] / 2,
                "y": data["top"][i] + data["height"][i] / 2,
            })
    return ocurrencias


def _producto_mas_cercano(anclas: list[dict], x: float, y: float) -> Optional[str]:
    if not anclas:
        return None
    return min(anclas, key=lambda a: (a["x"] - x) ** 2 + (a["y"] - y) ** 2)["codigo"]


def _detectar_ofertas_en_pagina(data: dict, ofertas_ref: list[sqlite3.Row]) -> list[dict]:
    """Para cada oferta de referencia que llega al umbral en el texto OCR
    de la pagina, ubica el/los producto(s) mas cercanos a donde aparecio
    la palabra clave. Si la pagina no tiene ningun 'cod. NNNNN' legible,
    devuelve la oferta sin producto asociado (degradacion visible a nivel
    de pagina, no una asignacion inventada a un producto al azar)."""
    texto_ocr = " ".join(t for t in data["text"] if t.strip())
    anclas = _anclas_producto(data)
    resultados = []
    for ref in ofertas_ref:
        ref_tokens = _tokens_significativos(ref["nombre_oferta"])
        if not ref_tokens:
            continue
        ocr_tokens = _tokens_significativos(texto_ocr)
        encontrados = [t for t in ref_tokens if _coincide_token(ocr_tokens, t)]
        score = len(encontrados) / len(ref_tokens)
        if score < UMBRAL_SCORE_OFERTA:
            continue
        posiciones = [p for t in encontrados for p in _ocurrencias_token(data, t)]
        if not posiciones:
            continue
        if not anclas:
            resultados.append({"oferta": ref["nombre_oferta"], "score": score, "producto_codigo": None})
            continue
        codigos_vistos = set()
        for pos in posiciones:
            codigo = _producto_mas_cercano(anclas, pos["x"], pos["y"])
            if codigo not in codigos_vistos:
                codigos_vistos.add(codigo)
                resultados.append({"oferta": ref["nombre_oferta"], "score": score, "producto_codigo": codigo})
    return resultados


# --- Homologacion de Venta Directa contra el catalogo real de Azzorti ---
# Retail homologa contra "azzorti_producto" (Excel real con una fila por
# producto). Venta Directa no tenia ningun catalogo Azzorti estructurado
# para eso - ahora se indexa (OCR) el catalogo PDF que se sube en
# Competencia con competidor="Azzorti" y se homologa por palabras clave
# contra el texto que quedo cerca de cada "COD." de ese catalogo.

_PRECIO_RE = re.compile(r"BS\.?\s*(\d+[.,]\d{2})")


def _seccion_de_pagina(data: dict, alto_pagina: float) -> Optional[str]:
    """El catalogo real marca cada seccion con un rotulo en mayusculas en
    el 8% inferior de la pagina (ej. 'FRAGANCIAS | 99', 'BELLEZA Y
    CUIDADO - HOGAR | 139') - se usa para no comparar una captura de
    Fragancias contra productos de Moda/Hogar que quedaron indexados en
    el mismo catalogo. Si el OCR no logro leer el rotulo de esa pagina en
    particular, devuelve None (la pagina no queda excluida, solo sin
    seccion identificada - ver _candidato_coincide_categoria)."""
    umbral_y = alto_pagina * 0.92
    palabras = []
    for i, texto in enumerate(data["text"]):
        t = texto.strip()
        if t and data["top"][i] >= umbral_y and t.isalpha() and t.isupper() and len(t) >= 3:
            palabras.append(t)
    return " ".join(palabras) if palabras else None


def _indexar_pagina_productos(data: dict, alto_pagina: float) -> list[dict]:
    """Para cada codigo de producto de la pagina (VD: 'cod. NNNNN' con el
    nombre/precio arriba; Retail: 'Nombre Ref.RNNNN' con el precio ABAJO -
    ver _anclas_producto_retail), junta el texto OCR que esta cerca."""
    anclas = (
        [dict(a, arriba=True) for a in _anclas_producto(data)]
        + [dict(a, arriba=False) for a in _anclas_producto_retail(data)]
    )
    seccion = _seccion_de_pagina(data, alto_pagina)
    resultados = []
    n = len(data["text"])
    for ancla in anclas:
        if ancla["arriba"]:
            y_min, y_max = ancla["y"] - 220, ancla["y"] + 60
        else:
            y_min, y_max = ancla["y"] - 20, ancla["y"] + 250
        cercanos = []
        for i in range(n):
            texto = data["text"][i].strip()
            if not texto:
                continue
            wx = data["left"][i] + data["width"][i] / 2
            wy = data["top"][i] + data["height"][i] / 2
            if y_min <= wy <= y_max and abs(wx - ancla["x"]) <= 350:
                cercanos.append(texto)
        texto_cercano = " ".join(cercanos)
        precio_m = _PRECIO_RE.search(_sin_acentos(texto_cercano.upper()))
        precio = float(precio_m.group(1).replace(",", ".")) if precio_m else None
        resultados.append({
            "producto_codigo": ancla["codigo"],
            "texto_cercano": texto_cercano[:500],
            "precio": precio,
            "y": ancla["y"],
            "seccion": seccion,
        })
    return resultados


# Categorias reales de Venta Directa (ver CompletarContextoScreen en la
# app) -> palabras que se esperan en el rotulo de seccion del catalogo
# Azzorti para esa categoria. Ninguna suposicion silenciosa: si la
# categoria no esta aqui o la pagina no tiene seccion identificada, el
# candidato NO se excluye (ver _candidato_coincide_categoria) - se
# prefiere mostrar de mas a arriesgarse a esconder el producto correcto.
_CATEGORIA_VD_A_SECCION = {
    "fragancias": ["FRAGANCIA", "FRAGANCIAS"],
    # "VOGUE" es la linea de maquillaje real de Azzorti en este catalogo -
    # esas paginas rotulan el pie con el nombre de la marca, no con la
    # palabra "Maquillaje" (a diferencia de otras secciones).
    "maquillaje": ["BELLEZA", "MAQUILLAJE", "ROSTRO", "VOGUE"],
    "rostro": ["BELLEZA", "ROSTRO", "MAQUILLAJE", "VOGUE"],
    "cabello": ["BELLEZA", "CABELLO", "CUIDADO"],
    "cuidado diario": ["BELLEZA", "CUIDADO"],
    "joyeria": ["JOYERIA", "ACCESORIOS"],
    "hogar": ["HOGAR"],
}


def _candidato_coincide_categoria(categoria_captura: str, seccion_pagina: Optional[str]) -> bool:
    esperadas = _CATEGORIA_VD_A_SECCION.get(_sin_acentos((categoria_captura or "").strip().lower()))
    if not esperadas or not seccion_pagina:
        return True
    seccion_norm = _sin_acentos(seccion_pagina.upper())
    return any(palabra in seccion_norm for palabra in esperadas)


_SINONIMOS_PRODUCTO = {
    "GAFAS": "LENTES",  # asi le dice el analista, asi le dice el catalogo real de Azzorti "lentes de sol"
}


def _variantes_palabra(token: str) -> set[str]:
    """Sinonimos conocidos + variantes de plural (POLVOS/POLVO) y genero
    gramatical (FEMENINA/FEMENINO) para comparar texto libre - separado de
    _coincide_token (que usa la deteccion de ofertas, ya afinada) para no
    arriesgar esa deteccion con un cambio pensado para otro problema.
    Devuelve un conjunto de variantes en vez de una sola forma normalizada
    porque encadenar "quitar plural" y "quitar genero" en un orden fijo
    fallaba (ej. POLVOS->POLVO por plural, pero POLVO->POLV por genero,
    terminaban en formas distintas igual)."""
    base = _SINONIMOS_PRODUCTO.get(token, token)
    variantes = {base}
    if base.isalpha():
        if len(base) >= 4 and base[-1] == "S":
            variantes.add(base[:-1])
        for v in list(variantes):
            if len(v) >= 5 and v[-1] in ("A", "O"):
                variantes.add(v[:-1])
    return variantes


def _coincide_token_texto_libre(ocr_tokens: set[str], ref_token: str) -> bool:
    if _coincide_token(ocr_tokens, ref_token):
        return True
    ref_variantes = _variantes_palabra(ref_token)
    return any(_variantes_palabra(t) & ref_variantes for t in ocr_tokens)


def _score_texto(texto_principal: str, texto_secundario: str, texto_catalogo: str) -> float:
    """Fraccion de palabras clave que aparecen en el texto OCR cercano a un
    producto del catalogo - separado en dos bloques con peso distinto:
    "principal" (categoria + descripcion, lo mas confiable que escribe el
    analista - un nombre corto de producto) pesa mas que "secundario"
    (caracteristicas + detalle, texto libre mas largo y con mas riesgo de
    palabras genericas). Antes pesaban igual, sin importar cuantas
    palabras tuviera cada uno (a peticion de Yohana: la descripcion debe
    influir mas que el texto libre de caracteristicas)."""
    tokens_catalogo = _tokens_significativos(texto_catalogo)

    def _fraccion(texto):
        tokens = _tokens_significativos(texto)
        if not tokens:
            return None
        encontrados = sum(1 for t in tokens if _coincide_token_texto_libre(tokens_catalogo, t))
        return encontrados / len(tokens)

    frac_principal = _fraccion(texto_principal)
    frac_secundario = _fraccion(texto_secundario)
    if frac_principal is None:
        return frac_secundario or 0.0
    if frac_secundario is None:
        return frac_principal
    combinado = 0.65 * frac_principal + 0.35 * frac_secundario
    # Si el usuario SI escribio caracteristicas especificas (2+ palabras
    # con contenido) y NINGUNA aparece en el candidato, es una señal fuerte
    # de que no es el mismo producto - se penaliza aunque el texto
    # principal tenga alguna palabra generica en comun (bug real: "horno
    # eléctrico" + caracteristicas "ahorro de luz y temporizador" contra
    # un secador de pelo solo compartia la palabra generica "eléctrico"
    # en el principal, nada en caracteristicas, e igual salia 22%).
    if frac_secundario == 0 and len(_tokens_significativos(texto_secundario)) >= 2:
        combinado *= 0.3
    return combinado


def _parsear_ofertas_referencia(contenido: bytes) -> list[dict]:
    """Lee 'Detalle de ofertas.xlsx': columnas Catalogo (competidor) y
    Nombre oferta, con la imagen del logotipo en la columna 'Visual'
    (imagen en celda, mismo mecanismo de rich-value que las fotos de
    productos estrella)."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb[wb.sheetnames[0]]
    encabezados = {}
    fila_encabezado = None
    for r in range(1, min(ws.max_row, 5) + 1):
        for c in range(1, ws.max_column + 1):
            valor = _texto(ws.cell(r, c).value).lower()
            if "catalogo" in valor:
                encabezados["competidor"] = c
                fila_encabezado = r
            elif "nombre" in valor and "oferta" in valor:
                encabezados["nombre_oferta"] = c
            elif "visual" in valor:
                encabezados["foto"] = c
    if fila_encabezado is None or "competidor" not in encabezados or "nombre_oferta" not in encabezados:
        raise ValueError("No se encontraron las columnas 'Catalogo' y 'Nombre oferta'.")
    filas = []
    for r in range(fila_encabezado + 1, ws.max_row + 1):
        competidor = _texto(ws.cell(r, encabezados["competidor"]).value)
        nombre_oferta = _texto(ws.cell(r, encabezados["nombre_oferta"]).value)
        if not competidor or not nombre_oferta:
            continue
        filas.append({
            "_fila_excel": r,
            "_col_foto": encabezados.get("foto"),
            "competidor": competidor,
            "nombre_oferta": nombre_oferta,
        })
    return filas


@app.post("/ofertas/importar")
async def importar_ofertas_referencia(archivo: UploadFile = File(...)):
    """Importa el set de referencia de logotipos de oferta desde el Excel
    real de Mercadeo ('Detalle de ofertas.xlsx'). Re-importar el mismo
    archivo actualiza (no duplica), por el UNIQUE(competidor, nombre_oferta)."""
    contenido = await archivo.read()
    try:
        filas = _parsear_ofertas_referencia(contenido)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el Excel: {e}")
    if not filas:
        raise HTTPException(400, "No se encontraron filas con Catalogo y Nombre oferta.")
    fotos_por_celda = _extraer_fotos_por_celda(contenido)

    def _guardar(foto_bytes: bytes, prefijo: str) -> str:
        extension = ".jpg" if foto_bytes[:3] == b"\xff\xd8\xff" else ".png"
        slug = re.sub(r"[^a-z0-9]+", "-", prefijo.lower()).strip("-")
        nombre = f"{slug}{extension}"
        (OFERTA_FOTOS_DIR / nombre).write_bytes(foto_bytes)
        return nombre

    con_foto = 0
    with get_conn() as conn:
        for f in filas:
            foto_bytes = fotos_por_celda.get((f["_fila_excel"], f["_col_foto"])) if f["_col_foto"] else None
            foto_nombre = None
            if foto_bytes:
                foto_nombre = _guardar(foto_bytes, f"{f['competidor']}-{f['nombre_oferta']}")
                con_foto += 1
            conn.execute(
                """INSERT INTO oferta_referencia (competidor, nombre_oferta, foto, actualizado_en)
                VALUES (?, ?, ?, ?)
                ON CONFLICT (competidor, nombre_oferta) DO UPDATE SET
                    foto=COALESCE(excluded.foto, oferta_referencia.foto),
                    actualizado_en=excluded.actualizado_en""",
                (f["competidor"], f["nombre_oferta"], foto_nombre, datetime.now(timezone.utc).isoformat()),
            )
    return {
        "mensaje": f"{len(filas)} ofertas de referencia importadas/actualizadas ({con_foto} con logotipo).",
        "total": len(filas),
    }


@app.get("/ofertas")
def listar_ofertas_referencia(request: Request):
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM oferta_referencia ORDER BY competidor, nombre_oferta").fetchall()
    return [
        {**dict(r), "foto_url": f"{request.base_url}static/ofertas_fotos/{r['foto']}" if r["foto"] else None}
        for r in rows
    ]


@app.post("/catalogos/{catalogo_id}/detectar-ofertas")
def detectar_ofertas_en_catalogo(catalogo_id: int):
    """Escanea (OCR) cada pagina del catalogo PDF y la compara contra las
    ofertas de referencia del mismo competidor. Es sincrono y puede tardar
    (1-3 seg por pagina) - pensado para un boton 'Detectar ofertas' en el
    panel de Competencia, no para correr automaticamente en cada carga."""
    with get_conn() as conn:
        catalogo = conn.execute("SELECT * FROM catalogo_competidor WHERE id = ?", (catalogo_id,)).fetchone()
        if not catalogo:
            raise HTTPException(404, "Catálogo no encontrado")
        ruta = CATALOGOS_DIR / catalogo["archivo"]
        if not ruta.exists():
            raise HTTPException(404, "El archivo del catálogo ya no está en el servidor.")
        if ruta.suffix.lower() != ".pdf":
            raise HTTPException(
                400,
                "Este catálogo no es un PDF (es una planilla) - las planillas de "
                "esta fuente no traen banda de oferta como imagen, solo la foto "
                "del producto, así que no hay nada que buscar aquí.",
            )
        ofertas_ref = conn.execute(
            "SELECT * FROM oferta_referencia WHERE LOWER(TRIM(competidor)) = LOWER(TRIM(?))",
            (catalogo["competidor"],),
        ).fetchall()
        if not ofertas_ref:
            raise HTTPException(
                400,
                f"No hay ofertas de referencia importadas para '{catalogo['competidor']}' "
                "- importa primero el Excel de ofertas (POST /ofertas/importar).",
            )
        doc = fitz.open(ruta)
        conn.execute("DELETE FROM catalogo_oferta WHERE catalogo_id = ?", (catalogo_id,))
        con_oferta = 0
        con_producto_identificado = 0
        ahora = datetime.now(timezone.utc).isoformat()
        for pno in range(doc.page_count):
            imagenes = doc.get_page_images(pno)
            if not imagenes:
                continue
            # si hay varias imagenes en la pagina se usa la mas grande
            # (algunas paginas traen un logo de marca chico ademas del
            # grafico principal) - se OCRea esa, no todas, por tiempo.
            xref = max(imagenes, key=lambda im: im[2] * im[3])[0]
            pix = fitz.Pixmap(doc, xref)
            if pix.colorspace is None or pix.colorspace.n != 3:
                pix = fitz.Pixmap(fitz.csRGB, pix)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            data = pytesseract.image_to_data(img, lang="spa", output_type=pytesseract.Output.DICT)
            texto_ocr = " ".join(t for t in data["text"] if t.strip())
            detecciones = _detectar_ofertas_en_pagina(data, ofertas_ref)
            if detecciones:
                con_oferta += 1
            for d in detecciones:
                if d["producto_codigo"]:
                    con_producto_identificado += 1
                conn.execute(
                    """INSERT INTO catalogo_oferta
                    (catalogo_id, pagina, producto_codigo, oferta_detectada, score, texto_ocr, creado_en)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (catalogo_id, pagina, oferta_detectada, producto_codigo) DO UPDATE SET
                        score=excluded.score, texto_ocr=excluded.texto_ocr, creado_en=excluded.creado_en""",
                    (catalogo_id, pno + 1, d["producto_codigo"], d["oferta"], round(d["score"], 2), texto_ocr[:2000], ahora),
                )
        conn.commit()
    return {
        "mensaje": f"{doc.page_count} páginas analizadas, {con_oferta} con oferta identificada "
                   f"({con_producto_identificado} con producto específico ubicado).",
        "paginas_analizadas": doc.page_count,
        "paginas_con_oferta": con_oferta,
        "con_producto_identificado": con_producto_identificado,
    }


@app.get("/catalogos/{catalogo_id}/ofertas")
def listar_ofertas_de_catalogo(catalogo_id: int):
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT pagina, producto_codigo, oferta_detectada, score FROM catalogo_oferta "
            "WHERE catalogo_id = ? ORDER BY pagina",
            (catalogo_id,),
        ).fetchall()
    return [dict(r) for r in rows]


@app.post("/catalogos/{catalogo_id}/indexar-productos")
def indexar_productos_de_catalogo(catalogo_id: int):
    """Lee (OCR) cada pagina del catalogo PDF y guarda que codigo de
    producto aparecio cerca de que texto/precio - para poder homologar
    Venta Directa contra este catalogo (ver /capturas/{id}/homologacion/
    sugerencias). Pensado sobre todo para el catalogo de competidor=
    'Azzorti', pero no se restringe a el por si mas adelante sirve para
    comparar tambien contra el catalogo de un competidor. Sincrono y
    puede tardar varios minutos en catalogos largos (renderiza cada
    pagina completa, no solo la imagen mas grande, porque este catalogo
    combina texto vectorial e imagenes en la misma pagina)."""
    with get_conn() as conn:
        catalogo = conn.execute("SELECT * FROM catalogo_competidor WHERE id = ?", (catalogo_id,)).fetchone()
        if not catalogo:
            raise HTTPException(404, "Catálogo no encontrado")
        ruta = CATALOGOS_DIR / catalogo["archivo"]
        if not ruta.exists():
            raise HTTPException(404, "El archivo del catálogo ya no está en el servidor.")
        if ruta.suffix.lower() != ".pdf":
            raise HTTPException(400, "Este catálogo no es un PDF - no hay página que renderizar.")
        doc = fitz.open(ruta)
        conn.execute("DELETE FROM catalogo_producto WHERE catalogo_id = ?", (catalogo_id,))
        ahora = datetime.now(timezone.utc).isoformat()
        total_productos = 0
        for pno in range(doc.page_count):
            pix = doc[pno].get_pixmap(dpi=150)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            data = pytesseract.image_to_data(img, lang="spa", output_type=pytesseract.Output.DICT)
            productos_pagina = _indexar_pagina_productos(data, pix.height)
            if productos_pagina:
                # Recorte por producto, no la pagina completa: varios
                # productos pueden compartir una misma pagina (ej. lentes +
                # reloj) y mostrar la pagina entera confundia al analista al
                # homologar (aparecia un producto que no tenia nada que ver).
                # Se divide la pagina en franjas horizontales usando el
                # punto medio entre las posiciones Y de codigos consecutivos
                # - es una aproximacion (asume que los productos de una
                # misma pagina se apilan verticalmente, que es el diseno
                # real de este catalogo), no una deteccion exacta del
                # recuadro de cada producto.
                ordenados = sorted(productos_pagina, key=lambda p: p["y"])
                for i, p in enumerate(ordenados):
                    arriba = 0 if i == 0 else (ordenados[i - 1]["y"] + p["y"]) / 2
                    abajo = pix.height if i == len(ordenados) - 1 else (p["y"] + ordenados[i + 1]["y"]) / 2
                    recorte = img.crop((0, max(0, arriba - 10), pix.width, min(pix.height, abajo + 10)))
                    recorte.save(CATALOGO_PAGINAS_DIR / f"{catalogo_id}_{pno + 1}_{p['producto_codigo']}.png")
            for p in productos_pagina:
                total_productos += 1
                conn.execute(
                    """INSERT INTO catalogo_producto
                    (catalogo_id, pagina, producto_codigo, texto_cercano, precio, seccion, creado_en)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (catalogo_id, pagina, producto_codigo) DO UPDATE SET
                        texto_cercano=excluded.texto_cercano, precio=excluded.precio,
                        seccion=excluded.seccion, creado_en=excluded.creado_en""",
                    (catalogo_id, pno + 1, p["producto_codigo"], p["texto_cercano"], p["precio"], p["seccion"], ahora),
                )
        conn.commit()
    return {
        "mensaje": f"{doc.page_count} páginas analizadas, {total_productos} productos indexados.",
        "paginas_analizadas": doc.page_count,
        "total_productos": total_productos,
    }


@app.get("/catalogos/{catalogo_id}/productos-resumen")
def resumen_productos_de_catalogo(catalogo_id: int):
    with get_conn() as conn:
        total = conn.execute(
            "SELECT COUNT(*) FROM catalogo_producto WHERE catalogo_id = ?", (catalogo_id,)
        ).fetchone()[0]
        con_precio = conn.execute(
            "SELECT COUNT(*) FROM catalogo_producto WHERE catalogo_id = ? AND precio IS NOT NULL",
            (catalogo_id,),
        ).fetchone()[0]
    return {"total_productos": total, "con_precio": con_precio}
